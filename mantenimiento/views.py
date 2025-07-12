import openpyxl
from django.shortcuts import render
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.http import JsonResponse
from .models import Patrono, Promocion, TargetPromocion, Agencias

def is_superuser(user):
    return user.is_superuser

@user_passes_test(is_superuser)
def moduloMantenimiento(request):
    """View for the maintenance module - only accessible by superusers"""
    context = {
        'patronos': Patrono.objects.all(),
        'promociones': Promocion.objects.all(),
        'target_promociones': TargetPromocion.objects.all(),
        'agencias': Agencias.objects.all(),
    }
    return render(request, "mantenimiento/moduloMantenimiento.html", context)

@user_passes_test(is_superuser)
def cargaAgencias(request):
    """Handle bulk upload of agencias from Excel file"""
    if request.method == "POST":
        file = request.FILES.get("file")
        
        if not file:
            return JsonResponse({
                'success': False,
                'error': 'No se proporcionó ningún archivo.'
            })
        
        if not file.name.endswith(".xlsx"):
            return JsonResponse({
                'success': False,
                'error': 'Por favor suba un archivo Excel válido (.xlsx).'
            })

        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_rows = []

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Skip header row and process data
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if len(row) < 2:
                        error_rows.append((idx, "Fila con menos de 2 columnas"))
                        continue
                    
                    secuencia, razon_social = row[0], row[1]
                    
                    # Validate data
                    if not secuencia or not razon_social:
                        error_rows.append((idx, "Secuencia o razón social vacía"))
                        continue
                    
                    # Convert secuencia to integer
                    try:
                        secuencia = int(secuencia)
                    except (ValueError, TypeError):
                        error_rows.append((idx, f"Secuencia '{secuencia}' no es un número válido"))
                        continue
                    
                    # Check for duplicates based on secuencia
                    existing_agencia = Agencias.objects.filter(secuencia=secuencia).first()
                    
                    if existing_agencia:
                        # Update existing record
                        existing_agencia.razon_social = str(razon_social).strip()
                        existing_agencia.save()
                        updated_count += 1
                    else:
                        # Create new record
                        Agencias.objects.create(
                            secuencia=secuencia,
                            razon_social=str(razon_social).strip()
                        )
                        created_count += 1
                        
                except Exception as e:
                    error_rows.append((idx, str(e)))

            # Prepare response
            response_data = {
                'success': True,
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'errors': error_rows
            }
            
            if error_rows:
                response_data['error_summary'] = f"Se encontraron errores en {len(error_rows)} filas."
            
            return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al procesar el archivo: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Método no permitido.'
    })

def cargaPatronos(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Please upload a valid .xlsx file.")
            return render(request, "mantenimiento/cargaPatronos.html")

        error_rows = []
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                try:
                    codigo, descripcion, agrupador, selectDescuento, porServDesc, montoFijo, disketCentral = row
                    Patrono.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            "descripcion": descripcion,
                            "agrupador": agrupador,
                            "selectDescuento": selectDescuento,
                            "porServDesc": porServDesc if porServDesc else None,
                            "montoFijo": montoFijo if montoFijo else None,
                            "disketCentral": disketCentral,
                        },
                    )
                except Exception as e:
                    error_rows.append((idx, str(e)))

            if error_rows:
                error_summary = "\n".join([f"Row {row}: {error}" for row, error in error_rows])
                messages.warning(request, f"Some rows were skipped due to errors:\n{error_summary}")
            else:
                messages.success(request, "Patrono list uploaded successfully.")
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")

    patronos = Patrono.objects.all()
    return render(request, "mantenimiento/cargaPatronos.html", {"patronos": patronos})