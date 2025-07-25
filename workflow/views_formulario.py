from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from .models import ClienteEntrevista, ReferenciaPersonal, ReferenciaComercial, OtroIngreso
from .forms import (
    ClienteEntrevistaForm,
    ReferenciaPersonalFormSet,
    ReferenciaComercialFormSet,
    OtroIngresoFormSet,
)
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import csv
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


def entrevista_cliente_view(request):
    """Vista principal del formulario de entrevista del cliente"""
    steps_info = [
        ("Datos Generales", "fa-user"),
        ("Dirección", "fa-location-dot"),
        ("Cónyuge", "fa-people-roof"),
        ("Laboral", "fa-briefcase"),
        ("PEP", "fa-shield-halved"),
        ("Ref. Pers.", "fa-user-group"),
        ("Ref. Com.", "fa-building"),
        ("Bancarios", "fa-university"),
        ("Autorizaciones", "fa-circle-check"),
    ]

    if request.method == 'POST':
        form = ClienteEntrevistaForm(request.POST)
        referencias_formset = ReferenciaPersonalFormSet(
            request.POST,
            queryset=ReferenciaPersonal.objects.none(),
            prefix='referencias'
        )
        referencias_comerciales_formset = ReferenciaComercialFormSet(
            request.POST,
            queryset=ReferenciaComercial.objects.none(),
            prefix='referencias_comerciales'
        )
        otros_ingresos_formset = OtroIngresoFormSet(
            request.POST,
            queryset=OtroIngreso.objects.none(),
            prefix='otros_ingresos'
        )

        tiene_otros_ingresos = request.POST.get('tiene_otros_ingresos') == 'true'

        if (form.is_valid() and referencias_formset.is_valid() and
            referencias_comerciales_formset.is_valid() and otros_ingresos_formset.is_valid()):
            cliente = form.save(commit=False)
            
            # Asigna explícitamente todos los datos generales
            cliente.primer_nombre = form.cleaned_data.get('primer_nombre')
            cliente.segundo_nombre = form.cleaned_data.get('segundo_nombre')
            cliente.primer_apellido = form.cleaned_data.get('primer_apellido')
            cliente.segundo_apellido = form.cleaned_data.get('segundo_apellido')
            cliente.provincia_cedula = form.cleaned_data.get('provincia_cedula')
            cliente.tipo_letra = form.cleaned_data.get('tipo_letra')
            cliente.tomo_cedula = form.cleaned_data.get('tomo_cedula')
            cliente.partida_cedula = form.cleaned_data.get('partida_cedula')
            cliente.telefono = form.cleaned_data.get('telefono')
            cliente.email = form.cleaned_data.get('email')
            cliente.fecha_nacimiento = form.cleaned_data.get('fecha_nacimiento')
            cliente.sexo = form.cleaned_data.get('sexo')
            cliente.jubilado = form.cleaned_data.get('jubilado')
            cliente.nivel_academico = form.cleaned_data.get('nivel_academico')
            cliente.estado_civil = form.cleaned_data.get('estado_civil')
            cliente.no_dependientes = form.cleaned_data.get('no_dependientes')
            cliente.titulo = form.cleaned_data.get('titulo')
            cliente.salario = form.cleaned_data.get('salario')
            cliente.tipo_producto = form.cleaned_data.get('tipo_producto')
            cliente.oficial = form.cleaned_data.get('oficial')
            cliente.apellido_casada = form.cleaned_data.get('apellido_casada')
            cliente.peso = form.cleaned_data.get('peso')
            cliente.estatura = form.cleaned_data.get('estatura')
            cliente.nacionalidad = form.cleaned_data.get('nacionalidad')

            # Otros ingresos del modelo principal
            cliente.tipo_ingreso_1 = form.cleaned_data.get('tipo_ingreso_1')
            cliente.descripcion_ingreso_1 = form.cleaned_data.get('descripcion_ingreso_1')
            cliente.monto_ingreso_1 = form.cleaned_data.get('monto_ingreso_1')
            cliente.tipo_ingreso_2 = form.cleaned_data.get('tipo_ingreso_2')
            cliente.descripcion_ingreso_2 = form.cleaned_data.get('descripcion_ingreso_2')
            cliente.monto_ingreso_2 = form.cleaned_data.get('monto_ingreso_2')
            cliente.tipo_ingreso_3 = form.cleaned_data.get('tipo_ingreso_3')
            cliente.descripcion_ingreso_3 = form.cleaned_data.get('descripcion_ingreso_3')
            cliente.monto_ingreso_3 = form.cleaned_data.get('monto_ingreso_3')

            # Información laboral
            cliente.trabajo_direccion = form.cleaned_data.get('trabajo_direccion')
            cliente.trabajo_lugar = form.cleaned_data.get('trabajo_lugar')
            cliente.trabajo_cargo = form.cleaned_data.get('trabajo_cargo')
            cliente.tipo_trabajo = form.cleaned_data.get('tipo_trabajo')
            cliente.frecuencia_pago = form.cleaned_data.get('frecuencia_pago')
            cliente.tel_trabajo = form.cleaned_data.get('tel_trabajo')
            cliente.tel_ext = form.cleaned_data.get('tel_ext')
            cliente.origen_fondos = form.cleaned_data.get('origen_fondos')
            cliente.fecha_inicio_trabajo = form.cleaned_data.get('fecha_inicio_trabajo')

            cliente.save()

            # Guardar referencias personales
            referencias = referencias_formset.save(commit=False)
            for ref in referencias:
                ref.entrevista = cliente
                ref.save()
            for ref in referencias_formset.deleted_objects:
                ref.delete()

            # Guardar referencias comerciales
            comerciales = referencias_comerciales_formset.save(commit=False)
            for com in comerciales:
                com.entrevista = cliente
                com.save()
            for com in referencias_comerciales_formset.deleted_objects:
                com.delete()

            # Guardar otros ingresos si los hay
            if tiene_otros_ingresos:
                otros = otros_ingresos_formset.save(commit=False)
                for otro in otros:
                    otro.cliente = cliente
                    otro.save()
                for otro in otros_ingresos_formset.deleted_objects:
                    otro.delete()
                    
            return redirect('formulario_gracias')
        else:
            # Print errors for debugging
            if form.errors:
                print('Errores en el formulario principal:', form.errors)
            print('Errores en referencias personales:', referencias_formset.errors)
            print('Errores en referencias comerciales:', referencias_comerciales_formset.errors)
            print('Errores en otros ingresos:', otros_ingresos_formset.errors)

    else:
        form = ClienteEntrevistaForm()
        referencias_formset = ReferenciaPersonalFormSet(
            queryset=ReferenciaPersonal.objects.none(),
            prefix='referencias'
        )
        referencias_comerciales_formset = ReferenciaComercialFormSet(
            queryset=ReferenciaComercial.objects.none(),
            prefix='referencias_comerciales'
        )
        otros_ingresos_formset = OtroIngresoFormSet(
            queryset=OtroIngreso.objects.none(),
            prefix='otros_ingresos'
        )

    return render(request, 'formulario/entrevista_cliente.html', {
        'form': form,
        'referencias_formset': referencias_formset,
        'referencias_comerciales_formset': referencias_comerciales_formset,
        'otros_ingresos_formset': otros_ingresos_formset,
        'steps_info': steps_info,
        'messages': messages.get_messages(request)
    })


def entrevista_admin_view(request, entrevista_id):
    """Vista del formulario en modo administrador para completar entrevistas de clientes"""
    try:
        # Obtener la entrevista existente
        entrevista = get_object_or_404(ClienteEntrevista, id=entrevista_id)
        
        steps_info = [
            ("Datos Generales", "fa-user"),
            ("Dirección", "fa-location-dot"),
            ("Cónyuge", "fa-people-roof"),
            ("Laboral", "fa-briefcase"),
            ("PEP", "fa-shield-halved"),
            ("Ref. Pers.", "fa-user-group"),
            ("Ref. Com.", "fa-building"),
            ("Bancarios", "fa-university"),
            ("Autorizaciones", "fa-circle-check"),
        ]

        if request.method == 'POST':
            # Formulario con datos existentes para edición
            form = ClienteEntrevistaForm(request.POST, instance=entrevista)
            
            # Cargar formsets con datos existentes
            referencias_formset = ReferenciaPersonalFormSet(
                request.POST,
                queryset=entrevista.referencias_personales.all(),
                prefix='referencias'
            )
            referencias_comerciales_formset = ReferenciaComercialFormSet(
                request.POST,
                queryset=entrevista.referencias_comerciales.all(),
                prefix='referencias_comerciales'
            )
            otros_ingresos_formset = OtroIngresoFormSet(
                request.POST,
                queryset=entrevista.otros_ingresos.all(),
                prefix='otros_ingresos'
            )

            tiene_otros_ingresos = request.POST.get('tiene_otros_ingresos') == 'true'

            # Validación relajada para modo admin (control de calidad)
            if form.is_valid():  # Solo validamos el form principal, formsets opcionales
                try:
                    # Guardar entrevista principal
                    entrevista_guardada = form.save()
                    
                    # Procesar referencias personales (si las hay)
                    if referencias_formset.is_valid():
                        referencias_formset.instance = entrevista_guardada
                        referencias_formset.save()
                    
                    # Procesar referencias comerciales (si las hay)
                    if referencias_comerciales_formset.is_valid():
                        referencias_comerciales_formset.instance = entrevista_guardada
                        referencias_comerciales_formset.save()
                    
                    # Procesar otros ingresos (si los hay)
                    if otros_ingresos_formset.is_valid():
                        otros_ingresos_formset.instance = entrevista_guardada
                        otros_ingresos_formset.save()
                    
                    # Marcar como completada por admin
                    entrevista_guardada.completada_por_admin = True
                    entrevista_guardada.fecha_completada_admin = datetime.datetime.now()
                    entrevista_guardada.save()
                    
                    messages.success(request, 
                        f'✅ Entrevista de {entrevista_guardada.primer_nombre} {entrevista_guardada.primer_apellido} '
                        f'completada exitosamente en MODO ADMINISTRADOR. Lista para generar PDF/Excel.')
                    
                    # Redirigir a la lista de entrevistas
                    return redirect('lista_entrevistas')
                    
                except Exception as e:
                    messages.error(request, f'Error al guardar en modo admin: {str(e)}')
            else:
                messages.warning(request, 'Revise los campos marcados en rojo. En modo admin, algunos campos pueden quedar vacíos.')

        else:
            # Cargar formulario con datos existentes
            form = ClienteEntrevistaForm(instance=entrevista)
            
            # Cargar formsets con datos existentes
            referencias_formset = ReferenciaPersonalFormSet(
                queryset=entrevista.referencias_personales.all(),
                prefix='referencias'
            )
            referencias_comerciales_formset = ReferenciaComercialFormSet(
                queryset=entrevista.referencias_comerciales.all(),
                prefix='referencias_comerciales'
            )
            otros_ingresos_formset = OtroIngresoFormSet(
                queryset=entrevista.otros_ingresos.all(),
                prefix='otros_ingresos'
            )

        return render(request, 'formulario/entrevista_admin.html', {
            'form': form,
            'entrevista': entrevista,
            'referencias_formset': referencias_formset,
            'referencias_comerciales_formset': referencias_comerciales_formset,
            'otros_ingresos_formset': otros_ingresos_formset,
            'steps_info': steps_info,
            'modo_admin': True,
            'messages': messages.get_messages(request)
        })
        
    except Exception as e:
        messages.error(request, f'Error al acceder al modo administrador: {str(e)}')
        return redirect('lista_entrevistas')


def gracias(request):
    """Vista de agradecimiento después de completar el formulario"""
    return render(request, 'formulario/gracias.html')


def lista_entrevistas(request):
    """Vista que muestra la lista de todas las entrevistas"""
    entrevistas = ClienteEntrevista.objects.values(
        'id', 'primer_nombre', 'primer_apellido', 'email', 'telefono', 'tipo_producto', 'oficial',
        'provincia_cedula', 'tipo_letra', 'tomo_cedula', 'partida_cedula', 'fecha_entrevista'
    )
    total_entrevistas = entrevistas.count()
    return render(request, 'formulario/lista_entrevistas.html', {
        'entrevistas': entrevistas,
        'total_entrevistas': total_entrevistas
    })


def descargar_entrevistas_excel(request):
    """Descarga todas las entrevistas en formato Excel"""
    entrevistas = ClienteEntrevista.objects.all()

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    # Hoja de entrevistas principales
    ws_entrevistas = wb.create_sheet(title="Entrevistas")
    campos_excluir = [
        'tipo_ingreso_1', 'descripcion_ingreso_1', 'monto_ingreso_1',
        'tipo_ingreso_2', 'descripcion_ingreso_2', 'monto_ingreso_2',
        'tipo_ingreso_3', 'descripcion_ingreso_3', 'monto_ingreso_3'
    ]
    campos_entrevista = [
        field for field in ClienteEntrevista._meta.fields
        if field.name != 'lugar_nacimiento' and field.name not in campos_excluir
    ]

    encabezados = [
        field.verbose_name if field.name in ['peso', 'estatura'] else field.verbose_name.title() if field.name in ['conyuge_cargo', 'conyuge_ingreso'] else field.name
        for field in campos_entrevista
    ]
    if 'nacionalidad' not in [field.name for field in campos_entrevista]:
        encabezados.append('nacionalidad')
    if "Peso (lb)" not in encabezados:
        encabezados.append("Peso (lb)")
    if "Estatura (m)" not in encabezados:
        encabezados.append("Estatura (m)")
    ws_entrevistas.append(encabezados)

    for entrevista in entrevistas:
        fila = []
        for field in campos_entrevista:
            valor = getattr(entrevista, field.name, "")
            if field.name in ['peso', 'estatura']:
                valor = str(valor) if valor is not None else ""
            elif field.name in ['autoriza_apc', 'acepta_datos', 'es_beneficiario_final', 'es_pep', 'es_familiar_pep']:
                valor = "Sí" if valor else "No"
            elif field.name in ['conyuge_ingreso'] and valor is not None:
                try:
                    valor = float(valor)
                except Exception:
                    pass
            elif valor is not None and hasattr(valor, 'normalize'):
                valor = str(valor)
            if isinstance(valor, datetime.datetime) and valor.tzinfo is not None:
                valor = valor.replace(tzinfo=None)
            fila.append(valor if valor is not None else "")
        
        if 'nacionalidad' not in [field.name for field in campos_entrevista]:
            fila.append(getattr(entrevista, 'nacionalidad', ''))
        if 'peso' not in [field.name for field in campos_entrevista]:
            fila.append(str(getattr(entrevista, 'peso', '')) if getattr(entrevista, 'peso', None) is not None else "")
        if 'estatura' not in [field.name for field in campos_entrevista]:
            fila.append(str(getattr(entrevista, 'estatura', '')) if getattr(entrevista, 'estatura', None) is not None else "")
        ws_entrevistas.append(fila)

    # Hoja de referencias personales
    ws_ref_personales = wb.create_sheet(title="Referencias Personales")
    campos_ref_personal = ['entrevista_id', 'nombre', 'telefono', 'relacion', 'direccion']
    ws_ref_personales.append(campos_ref_personal)

    for entrevista in entrevistas:
        for ref in entrevista.referencias_personales.all():
            fila = [entrevista.id, ref.nombre, ref.telefono, ref.relacion, ref.direccion]
            ws_ref_personales.append(fila)

    # Hoja de referencias comerciales
    ws_ref_comerciales = wb.create_sheet(title="Referencias Comerciales")
    campos_ref_comercial = ['entrevista_id', 'tipo', 'nombre', 'actividad', 'telefono', 'celular', 'saldo']
    ws_ref_comerciales.append(campos_ref_comercial)

    for entrevista in entrevistas:
        for ref in entrevista.referencias_comerciales.all():
            fila = [entrevista.id, ref.tipo, ref.nombre, ref.actividad, ref.telefono, ref.celular, ref.saldo]
            ws_ref_comerciales.append(fila)

    # Hoja de otros ingresos
    ws_otros_ingresos = wb.create_sheet(title="Otros Ingresos")
    campos_otros_ingresos = ['cliente_id', 'tipo_ingreso', 'fuente', 'monto']
    ws_otros_ingresos.append(campos_otros_ingresos)

    for entrevista in entrevistas:
        for ingreso in entrevista.otros_ingresos.all():
            fila = [
                entrevista.id,
                ingreso.tipo_ingreso,
                ingreso.fuente,
                ingreso.monto
            ]
            ws_otros_ingresos.append(fila)

    # Ajustar ancho de columnas
    for ws in [ws_entrevistas, ws_ref_personales, ws_ref_comerciales, ws_otros_ingresos]:
        for col_num, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=entrevistas_completas.xlsx'
    wb.save(response)
    return response


def descargar_entrevista_excel(request, entrevista_id):
    """Descarga una entrevista específica en formato CSV"""
    entrevista = ClienteEntrevista.objects.get(pk=entrevista_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="entrevista_{entrevista_id}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Campo', 'Valor'])

    for field in entrevista._meta.fields:
        writer.writerow([field.verbose_name, getattr(entrevista, field.name)])

    writer.writerow([])
    writer.writerow(['Referencias Personales'])
    for ref in entrevista.referencias_personales.all():
        writer.writerow([ref.nombre, ref.direccion, ref.relacion, '', '', '', ''])

    writer.writerow([])
    writer.writerow(['Referencias Comerciales'])
    for ref in entrevista.referencias_comerciales.all():
        writer.writerow([ref.nombre, ref.tipo, ref.actividad, ref.telefono, ref.celular, ref.saldo])

    writer.writerow([])
    writer.writerow(['Otros Ingresos'])
    for ingreso in entrevista.otros_ingresos.all():
        writer.writerow([ingreso.fuente, '', ingreso.monto])

    return response


def descargar_entrevista_pdf(request, entrevista_id):
    """Descarga una entrevista específica en formato PDF COMPACTO - 2-3 páginas máximo! 🌟"""
    try:
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.platypus import PageBreak, KeepTogether
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        
        # Obtener la entrevista completa
        entrevista = get_object_or_404(ClienteEntrevista, id=entrevista_id)
        
        print(f"Generando PDF COMPACTO para: {entrevista.primer_nombre} {entrevista.primer_apellido}")
        
        # Crear buffer de memoria
        buffer = BytesIO()
        
        # Colores corporativos Pacífico
        pacifico_blue = HexColor('#003366')
        pacifico_green = HexColor('#00A651')
        pacifico_light_blue = HexColor('#E6F3FF')
        pacifico_light_green = HexColor('#E8F5E8')
        
        # Crear documento PDF con márgenes optimizados pero legibles
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter, 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch,
            leftMargin=0.65*inch,
            rightMargin=0.65*inch
        )
        
        # Estilos optimizados con espaciado dinámico
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'OptimizedTitle',
            parent=styles['Title'],
            fontSize=20,
            spaceAfter=8,
            textColor=pacifico_blue,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        section_header_style = ParagraphStyle(
            'OptimizedSectionHeader',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=6,
            spaceBefore=10,
            textColor=pacifico_blue,
            backColor=pacifico_light_blue,
            leftIndent=8,
            rightIndent=8,
            borderPadding=4,
            fontName='Helvetica-Bold'
        )
        
        story = []
        
        # =================== ENCABEZADO OPTIMIZADO ===================
        header_data = [
            ['FINANCIERA PACÍFICO S.A. - FORMULARIO DE ENTREVISTA', f'Doc: PAC-{entrevista_id:06d}'],
            [f'Fecha: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}', f'Cliente: {entrevista.primer_nombre} {entrevista.primer_apellido}']
        ]
        
        header_table = Table(header_data, colWidths=[4.2*inch, 2.8*inch])
        header_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 13),
            ('FONTSIZE', (1, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, 0), pacifico_blue),
            ('TEXTCOLOR', (1, 0), (-1, -1), pacifico_green),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (-1, -1), pacifico_light_green),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 12))
        
        # =================== SECCIÓN I: IDENTIFICACIÓN PERSONAL OPTIMIZADA ===================
        story.append(Paragraph("I. IDENTIFICACIÓN PERSONAL", section_header_style))
        
        # Función para ajustar texto según longitud
        def format_field_value(value, max_length=25):
            if not value:
                return '_' * min(max_length, 20)
            str_value = str(value)
            if len(str_value) > max_length:
                return str_value[:max_length-3] + '...'
            return str_value + '_' * max(0, max_length - len(str_value))
        
        # Dividir en dos columnas con espaciado dinámico
        datos_personales_col1 = [
            ['👤 Primer Nombre:', format_field_value(entrevista.primer_nombre, 20)],
            ['👤 Segundo Nombre:', format_field_value(entrevista.segundo_nombre, 20)],
            ['👤 Primer Apellido:', format_field_value(entrevista.primer_apellido, 20)],
            ['👤 Segundo Apellido:', format_field_value(entrevista.segundo_apellido, 20)],
            ['💍 Apellido Casada:', format_field_value(entrevista.apellido_casada, 20)],
            ['🆔 Cédula:', f"{entrevista.provincia_cedula or '_'}-{entrevista.tipo_letra or '_'}-{entrevista.tomo_cedula or '___'}-{entrevista.partida_cedula or '___'}"],
            ['🎂 Fec. Nacimiento:', entrevista.fecha_nacimiento.strftime('%d/%m/%Y') if entrevista.fecha_nacimiento else '__/__/____'],
            ['⚧️ Sexo:', entrevista.sexo or '☐M ☐F'],
        ]
        
        datos_personales_col2 = [
            ['💒 Estado Civil:', format_field_value(entrevista.estado_civil, 18)],
            ['🌍 Nacionalidad:', format_field_value(getattr(entrevista, 'nacionalidad', ''), 18)],
            ['🎓 Nivel Académico:', format_field_value(entrevista.nivel_academico, 18)],
            ['👶 Dependientes:', str(entrevista.no_dependientes) if entrevista.no_dependientes else '__'],
            ['⚖️ Peso (lb):', str(entrevista.peso) if entrevista.peso else '___'],
            ['📏 Estatura (m):', str(entrevista.estatura) if entrevista.estatura else '___'],
            ['👴 Jubilado:', '✅Sí' if entrevista.jubilado else '❌No' if entrevista.jubilado is False else '☐Sí ☐No'],
            ['', ''],
        ]
        
        # Crear tabla de dos columnas con mejor espaciado
        tabla_datos = []
        for i in range(max(len(datos_personales_col1), len(datos_personales_col2))):
            row = []
            if i < len(datos_personales_col1):
                row.extend(datos_personales_col1[i])
            else:
                row.extend(['', ''])
            if i < len(datos_personales_col2):
                row.extend(datos_personales_col2[i])
            else:
                row.extend(['', ''])
            tabla_datos.append(row)
        
        tabla_personales = Table(tabla_datos, colWidths=[1.3*inch, 2.0*inch, 1.3*inch, 2.0*inch])
        tabla_personales.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('TEXTCOLOR', (2, 0), (2, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, pacifico_green),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_personales)
        story.append(Spacer(1, 10))
        
        # =================== SECCIÓN II: CONTACTO Y LABORAL OPTIMIZADAS ===================
        story.append(Paragraph("II. CONTACTO Y INFORMACIÓN LABORAL", section_header_style))
        
        # Obtener fecha de inicio de trabajo
        fecha_inicio = getattr(entrevista, 'fecha_inicio_trabajo', None)
        fecha_inicio_str = fecha_inicio.strftime('%d/%m/%Y') if fecha_inicio else '__/__/____'
        
        datos_contacto_laboral = [
            ['📱 Teléfono:', format_field_value(entrevista.telefono, 18), '🏢 Empresa:', format_field_value(getattr(entrevista, 'trabajo_lugar', ''), 18)],
            ['📧 Email:', format_field_value(entrevista.email, 18), '💼 Cargo:', format_field_value(entrevista.titulo, 18)],
            ['🏠 Dirección:', format_field_value(getattr(entrevista, 'direccion', ''), 18), '💰 Salario:', f"B/. {entrevista.salario:,.2f}" if entrevista.salario else 'B/. __________'],
            ['📍 Provincia:', format_field_value(getattr(entrevista, 'provincia_residencia', ''), 18), '📞 Tel. Trabajo:', format_field_value(getattr(entrevista, 'tel_trabajo', ''), 18)],
            ['', '', '📅 Fec. Inicio:', fecha_inicio_str],
        ]
        
        tabla_contacto_laboral = Table(datos_contacto_laboral, colWidths=[1.3*inch, 2.0*inch, 1.3*inch, 2.0*inch])
        tabla_contacto_laboral.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('TEXTCOLOR', (2, 0), (2, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, pacifico_green),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_contacto_laboral)
        story.append(Spacer(1, 10))
        
        # =================== SECCIÓN III: CÓNYUGE Y PEP OPTIMIZADAS ===================
        story.append(Paragraph("III. CÓNYUGE Y PERSONA EXPUESTA POLÍTICAMENTE", section_header_style))
        
        es_pep = '✅Sí' if getattr(entrevista, 'es_pep', None) else '❌No' if getattr(entrevista, 'es_pep', None) is False else '☐Sí ☐No'
        es_familiar_pep = '✅Sí' if getattr(entrevista, 'es_familiar_pep', None) else '❌No' if getattr(entrevista, 'es_familiar_pep', None) is False else '☐Sí ☐No'
        es_beneficiario = '✅Sí' if getattr(entrevista, 'es_beneficiario_final', None) else '❌No' if getattr(entrevista, 'es_beneficiario_final', None) is False else '☐Sí ☐No'
        
        datos_conyuge_pep = [
            ['👥 Nombre Cónyuge:', format_field_value(getattr(entrevista, 'conyuge_nombre', ''), 18), '🏛️ ¿Es PEP?:', es_pep],
            ['🏢 Trabajo Cónyuge:', format_field_value(getattr(entrevista, 'conyuge_trabajo', ''), 18), '👨‍👩‍👧‍👦 ¿Familiar PEP?:', es_familiar_pep],
            ['💰 Ingreso Cónyuge:', f"B/. {entrevista.conyuge_ingreso:,.2f}" if getattr(entrevista, 'conyuge_ingreso', None) else 'B/. __________', '🎯 ¿Beneficiario?:', es_beneficiario],
            ['📱 Tel. Cónyuge:', format_field_value(getattr(entrevista, 'conyuge_telefono', ''), 18), '📝 Explicación PEP:', format_field_value(getattr(entrevista, 'explicacion_pep', ''), 18)],
        ]
        
        tabla_conyuge_pep = Table(datos_conyuge_pep, colWidths=[1.3*inch, 2.0*inch, 1.3*inch, 2.0*inch])
        tabla_conyuge_pep.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('TEXTCOLOR', (2, 0), (2, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, pacifico_green),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 2),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_conyuge_pep)
        story.append(Spacer(1, 10))
        
        # =================== SECCIÓN IV: PRODUCTO Y AUTORIZACIONES OPTIMIZADAS ===================
        story.append(Paragraph("IV. PRODUCTO FINANCIERO Y AUTORIZACIONES", section_header_style))
        
        autoriza_apc = '✅Sí' if getattr(entrevista, 'autoriza_apc', None) else '❌No' if getattr(entrevista, 'autoriza_apc', None) is False else '☐Sí ☐No'
        acepta_datos = '✅Sí' if getattr(entrevista, 'acepta_datos', None) else '❌No' if getattr(entrevista, 'acepta_datos', None) is False else '☐Sí ☐No'
        
        datos_producto_auth = [
            ['🏦 Tipo Producto:', entrevista.tipo_producto or '☐Auto ☐Personal ☐Hipotecario', '🔍 Autoriza APC:', autoriza_apc],
            ['👔 Oficial:', format_field_value(entrevista.oficial, 18), '🛡️ Acepta Datos:', acepta_datos],
            ['📅 Fec. Entrevista:', entrevista.fecha_entrevista.strftime('%d/%m/%Y %H:%M') if entrevista.fecha_entrevista else '__/__/____ __:__', '📋 Verificaciones:', '☐Sí ☐No'],
            ['📋 Estado:', '✅EN PROCESO' if entrevista.primer_nombre else '⚠️PENDIENTE', '📞 Contacto Tel.:', '☐Sí ☐No'],
        ]
        
        tabla_producto_auth = Table(datos_producto_auth, colWidths=[1.3*inch, 2.0*inch, 1.3*inch, 2.0*inch])
        tabla_producto_auth.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('TEXTCOLOR', (2, 0), (2, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, pacifico_green),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 2),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_producto_auth)
        story.append(Spacer(1, 15))
        
        # =================== SECCIÓN DE FIRMAS OPTIMIZADA ===================
        story.append(Paragraph("FIRMAS Y VALIDACIÓN OFICIAL", section_header_style))
        story.append(Spacer(1, 12))
        
        # Tabla de firmas con mejor espaciado
        firmas_data = [
            ['_____________________________', '_____________________________'],
            ['FIRMA DEL CLIENTE', 'FIRMA DEL OFICIAL DE CRÉDITO'],
            [f"{entrevista.primer_nombre} {entrevista.primer_apellido}".upper(), (entrevista.oficial or 'OFICIAL ASIGNADO').upper()],
            [f"C.I.P. {entrevista.provincia_cedula or '__'}-{entrevista.tipo_letra or '__'}-{entrevista.tomo_cedula or '____'}-{entrevista.partida_cedula or '_____'}", 'FINANCIERA PACÍFICO S.A.'],
        ]
        
        firma_table = Table(firmas_data, colWidths=[3.5*inch, 3.5*inch])
        firma_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 1), (-1, 1), pacifico_blue),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 2), (-1, 2), 9),
            ('FONTSIZE', (0, 3), (-1, 3), 8),
            ('TEXTCOLOR', (0, 3), (-1, 3), HexColor('#666666')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.3, pacifico_green),
        ]))
        story.append(firma_table)
        story.append(Spacer(1, 12))
        
        # Footer optimizado
        fecha_actual = datetime.datetime.now()
        footer_data = [
            [f"📅 {fecha_actual.strftime('%d/%m/%Y')}", f"🕐 {fecha_actual.strftime('%H:%M hrs')}", f"📄 PAC-{entrevista_id:06d}", "© FINANCIERA PACÍFICO S.A."]
        ]
        
        footer_table = Table(footer_data, colWidths=[1.75*inch, 1.75*inch, 1.75*inch, 1.75*inch])
        footer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TEXTCOLOR', (0, 0), (-1, -1), pacifico_green),
            ('BACKGROUND', (0, 0), (-1, -1), pacifico_light_green),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(footer_table)
        
        # Generar PDF
        doc.build(story)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"FormularioOptimizado_PAC{entrevista_id:06d}_{entrevista.primer_nombre}_{entrevista.primer_apellido}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print("� PDF OPTIMIZADO generado exitosamente - Espaciado dinámico según caracteres! �")
        return response
        
    except Exception as e:
        print(f"Error en descargar_entrevista_pdf: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)


def descargar_excel_entrevistas(request):
    """Descargar todas las entrevistas en formato Excel"""
    try:
        entrevistas = ClienteEntrevista.objects.all()
        
        # Crear un nuevo libro de trabajo y una hoja
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Entrevistas"
        
        # Definir las cabeceras
        headers = [
            'ID', 'Primer Nombre', 'Segundo Nombre', 'Primer Apellido', 
            'Segundo Apellido', 'Apellido Casada', 'Cédula', 'Teléfono', 
            'Email', 'Fecha Entrevista', 'Oficial', 'Tipo Producto',
            'Salario', 'Estado Civil', 'Fecha Nacimiento'
        ]
        
        # Escribir las cabeceras
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            sheet[f'{column_letter}1'] = header
            sheet[f'{column_letter}1'].font = openpyxl.styles.Font(bold=True)
        
        # Escribir los datos
        for row_num, entrevista in enumerate(entrevistas, 2):
            cedula = f"{entrevista.provincia_cedula or ''}-{entrevista.tipo_letra or ''}-{entrevista.tomo_cedula or ''}-{entrevista.partida_cedula or ''}"
            
            data = [
                entrevista.id,
                entrevista.primer_nombre,
                entrevista.segundo_nombre,
                entrevista.primer_apellido,
                entrevista.segundo_apellido,
                entrevista.apellido_casada,
                cedula,
                entrevista.telefono,
                entrevista.email,
                entrevista.fecha_entrevista.strftime('%d/%m/%Y %H:%M') if entrevista.fecha_entrevista else '',
                entrevista.oficial,
                entrevista.tipo_producto,
                entrevista.salario,
                entrevista.estado_civil,
                entrevista.fecha_nacimiento.strftime('%d/%m/%Y') if entrevista.fecha_nacimiento else '',
            ]
            
            for col_num, value in enumerate(data, 1):
                column_letter = get_column_letter(col_num)
                sheet[f'{column_letter}{row_num}'] = value
        
        # Ajustar el ancho de las columnas
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        # Crear respuesta HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="entrevistas_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        return response
        
    except Exception as e:
        print(f"Error en descargar_excel_entrevistas: {str(e)}")
        return HttpResponse(f"Error al generar Excel: {str(e)}", status=500)
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=12,
            textColor=pacifico_green,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        section_header_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=8,
            spaceBefore=15,
            textColor=pacifico_blue,
            backColor=pacifico_light_blue,
            leftIndent=10,
            rightIndent=10,
            borderPadding=6,
            fontName='Helvetica-Bold'
        )
        
        subsection_style = ParagraphStyle(
            'Subsection',
            parent=styles['Heading3'],
            fontSize=12,
            spaceAfter=6,
            spaceBefore=8,
            textColor=pacifico_green,
            fontName='Helvetica-Bold'
        )
        
        story = []
        
        # =================== ENCABEZADO CORPORATIVO ===================
        # Logo simulado y header principal
        header_data = [
            ['', 'FINANCIERA PACÍFICO S.A.', ''],
            ['', 'FORMULARIO DE ENTREVISTA AL CLIENTE', ''],
            ['', f'Documento No. PDF-{entrevista_id:06d}', '']
        ]
        
        header_table = Table(header_data, colWidths=[1.5*inch, 4*inch, 1.5*inch])
        header_table.setStyle(TableStyle([
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, 0), 20),
            ('TEXTCOLOR', (1, 0), (1, 0), pacifico_blue),
            ('ALIGN', (1, 0), (1, 2), 'CENTER'),
            ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 1), (1, 1), 14),
            ('TEXTCOLOR', (1, 1), (1, 1), pacifico_green),
            ('FONTNAME', (1, 2), (1, 2), 'Helvetica'),
            ('FONTSIZE', (1, 2), (1, 2), 10),
            ('TEXTCOLOR', (1, 2), (1, 2), black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 15))
        
        # Línea decorativa
        line_data = [['_' * 100]]
        line_table = Table(line_data, colWidths=[7*inch])
        line_table.setStyle(TableStyle([
            ('TEXTCOLOR', (0, 0), (0, 0), pacifico_green),
            ('FONTSIZE', (0, 0), (0, 0), 12),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ]))
        story.append(line_table)
        story.append(Spacer(1, 20))
        
        # =================== INFORMACIÓN GENERAL DEL DOCUMENTO ===================
        fecha_actual = datetime.datetime.now()
        info_general = [
            ['FECHA DE GENERACIÓN:', fecha_actual.strftime('%d de %B de %Y - %H:%M hrs')],
            ['ID ÚNICO DEL CLIENTE:', f'PAC-{entrevista_id:06d}'],
            ['FECHA DE ENTREVISTA:', entrevista.fecha_entrevista.strftime('%d de %B de %Y - %H:%M hrs') if entrevista.fecha_entrevista else 'PENDIENTE DE PROGRAMAR'],
            ['ESTADO DEL FORMULARIO:', '✅ COMPLETADO' if entrevista.primer_nombre else '⚠️ EN PROCESO'],
        ]
        
        tabla_info = Table(info_general, colWidths=[2.2*inch, 4.8*inch])
        tabla_info.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green]),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_info)
        story.append(Spacer(1, 25))
        
        # =================== SECCIÓN I: IDENTIFICACIÓN PERSONAL ===================
        story.append(Paragraph("I. IDENTIFICACIÓN PERSONAL", section_header_style))
        
        datos_personales = [
            ['👤 PRIMER NOMBRE:', entrevista.primer_nombre or '________________________________'],
            ['👤 SEGUNDO NOMBRE:', entrevista.segundo_nombre or '________________________________'],
            ['👤 PRIMER APELLIDO:', entrevista.primer_apellido or '________________________________'],
            ['👤 SEGUNDO APELLIDO:', entrevista.segundo_apellido or '________________________________'],
            ['💍 APELLIDO DE CASADA:', entrevista.apellido_casada or '________________________________'],
            ['🆔 CÉDULA DE IDENTIDAD:', f"{entrevista.provincia_cedula or '__'}-{entrevista.tipo_letra or '__'}-{entrevista.tomo_cedula or '_____'}-{entrevista.partida_cedula or '_____'}"],
            ['🎂 FECHA DE NACIMIENTO:', entrevista.fecha_nacimiento.strftime('%d/%m/%Y') if entrevista.fecha_nacimiento else '__/__/____'],
            ['⚧️ SEXO:', entrevista.sexo or '☐ Masculino  ☐ Femenino'],
            ['💒 ESTADO CIVIL:', entrevista.estado_civil or '________________________________'],
            ['🌍 NACIONALIDAD:', getattr(entrevista, 'nacionalidad', '') or '________________________________'],
            ['🎓 NIVEL ACADÉMICO:', entrevista.nivel_academico or '________________________________'],
            ['👶 NO. DE DEPENDIENTES:', str(entrevista.no_dependientes) if entrevista.no_dependientes else '____'],
            ['⚖️ PESO (libras):', str(entrevista.peso) if entrevista.peso else '____'],
            ['📏 ESTATURA (metros):', str(entrevista.estatura) if entrevista.estatura else '____'],
            ['👴 ¿ES JUBILADO?:', '✅ SÍ' if entrevista.jubilado else '❌ NO' if entrevista.jubilado is False else '☐ SÍ  ☐ NO'],
        ]
        
        tabla_personales = Table(datos_personales, colWidths=[2.5*inch, 4.5*inch])
        tabla_personales.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_personales)
        story.append(Spacer(1, 20))
        
        # =================== SECCIÓN II: INFORMACIÓN DE CONTACTO ===================
        story.append(Paragraph("II. INFORMACIÓN DE CONTACTO", section_header_style))
        
        datos_contacto = [
            ['📱 TELÉFONO PERSONAL:', entrevista.telefono or '________________________________'],
            ['📧 CORREO ELECTRÓNICO:', entrevista.email or '________________________________'],
            ['🏠 DIRECCIÓN RESIDENCIAL:', getattr(entrevista, 'direccion', '') or '________________________________'],
            ['📍 DISTRITO/CORREGIMIENTO:', getattr(entrevista, 'distrito', '') or '________________________________'],
            ['🏙️ PROVINCIA:', getattr(entrevista, 'provincia_residencia', '') or '________________________________'],
        ]
        
        tabla_contacto = Table(datos_contacto, colWidths=[2.5*inch, 4.5*inch])
        tabla_contacto.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_contacto)
        story.append(Spacer(1, 20))
        
        # =================== SECCIÓN III: INFORMACIÓN LABORAL ===================
        story.append(Paragraph("III. INFORMACIÓN LABORAL Y FINANCIERA", section_header_style))
        
        # Obtener fecha de inicio de trabajo
        fecha_inicio = getattr(entrevista, 'fecha_inicio_trabajo', None)
        fecha_inicio_str = fecha_inicio.strftime('%d/%m/%Y') if fecha_inicio else '__/__/____'
        
        datos_laborales = [
            ['🏢 EMPRESA/EMPLEADOR:', getattr(entrevista, 'trabajo_lugar', '') or '________________________________'],
            ['📍 DIRECCIÓN DEL TRABAJO:', getattr(entrevista, 'trabajo_direccion', '') or '________________________________'],
            ['💼 CARGO/TÍTULO:', entrevista.titulo or '________________________________'],
            ['⚙️ TIPO DE TRABAJO:', getattr(entrevista, 'tipo_trabajo', '') or '☐ Público  ☐ Privado  ☐ Independiente'],
            ['💰 SALARIO MENSUAL:', f"B/. {entrevista.salario:,.2f}" if entrevista.salario else 'B/. ________________'],
            ['📅 FRECUENCIA DE PAGO:', getattr(entrevista, 'frecuencia_pago', '') or '☐ Quincenal  ☐ Mensual  ☐ Otro'],
            ['📞 TELÉFONO DEL TRABAJO:', getattr(entrevista, 'tel_trabajo', '') or '________________________________'],
            ['📟 EXTENSIÓN:', getattr(entrevista, 'tel_ext', '') or '________'],
            ['📆 FECHA INICIO TRABAJO:', fecha_inicio_str],
            ['💸 ORIGEN DE FONDOS:', getattr(entrevista, 'origen_fondos', '') or '☐ Local  ☐ Extranjero'],
        ]
        
        tabla_laborales = Table(datos_laborales, colWidths=[2.5*inch, 4.5*inch])
        tabla_laborales.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_laborales)
        story.append(Spacer(1, 20))
        
        # =================== NUEVA PÁGINA ===================
        story.append(PageBreak())
        
        # =================== SECCIÓN IV: INFORMACIÓN DEL CÓNYUGE ===================
        story.append(Paragraph("IV. INFORMACIÓN DEL CÓNYUGE/CONVIVIENTE", section_header_style))
        
        datos_conyuge = [
            ['👥 NOMBRE COMPLETO:', getattr(entrevista, 'conyuge_nombre', '') or '________________________________'],
            ['🏢 LUGAR DE TRABAJO:', getattr(entrevista, 'conyuge_trabajo', '') or '________________________________'],
            ['💼 CARGO/OCUPACIÓN:', getattr(entrevista, 'conyuge_cargo', '') or '________________________________'],
            ['💰 INGRESO MENSUAL:', f"B/. {entrevista.conyuge_ingreso:,.2f}" if getattr(entrevista, 'conyuge_ingreso', None) else 'B/. ________________'],
            ['📱 TELÉFONO DE CONTACTO:', getattr(entrevista, 'conyuge_telefono', '') or '________________________________'],
        ]
        
        tabla_conyuge = Table(datos_conyuge, colWidths=[2.5*inch, 4.5*inch])
        tabla_conyuge.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_conyuge)
        story.append(Spacer(1, 20))
        
        # =================== SECCIÓN V: PERSONA EXPUESTA POLÍTICAMENTE ===================
        story.append(Paragraph("V. PERSONA EXPUESTA POLÍTICAMENTE (PEP)", section_header_style))
        
        es_pep = '✅ SÍ' if getattr(entrevista, 'es_pep', None) else '❌ NO' if getattr(entrevista, 'es_pep', None) is False else '☐ SÍ  ☐ NO'
        es_familiar_pep = '✅ SÍ' if getattr(entrevista, 'es_familiar_pep', None) else '❌ NO' if getattr(entrevista, 'es_familiar_pep', None) is False else '☐ SÍ  ☐ NO'
        es_beneficiario = '✅ SÍ' if getattr(entrevista, 'es_beneficiario_final', None) else '❌ NO' if getattr(entrevista, 'es_beneficiario_final', None) is False else '☐ SÍ  ☐ NO'
        
        datos_pep = [
            ['🏛️ ¿ES USTED UNA PEP?:', es_pep],
            ['👨‍👩‍👧‍👦 ¿ES FAMILIAR DE UNA PEP?:', es_familiar_pep],
            ['🎯 ¿ES BENEFICIARIO FINAL?:', es_beneficiario],
            ['📝 EXPLICACIÓN/DETALLE PEP:', getattr(entrevista, 'explicacion_pep', '') or '________________________________'],
        ]
        
        tabla_pep = Table(datos_pep, colWidths=[2.5*inch, 4.5*inch])
        tabla_pep.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_pep)
        story.append(Spacer(1, 20))
        
        # =================== SECCIÓN VI: PRODUCTO FINANCIERO ===================
        story.append(Paragraph("VI. INFORMACIÓN DEL PRODUCTO FINANCIERO", section_header_style))
        
        datos_producto = [
            ['🏦 TIPO DE PRODUCTO:', entrevista.tipo_producto or '☐ Auto  ☐ Personal  ☐ Hipotecario  ☐ Otro'],
            ['👔 OFICIAL ASIGNADO:', entrevista.oficial or '________________________________'],
            ['📅 FECHA DE ENTREVISTA:', entrevista.fecha_entrevista.strftime('%d/%m/%Y a las %H:%M hrs') if entrevista.fecha_entrevista else '__/__/____ a las __:__ hrs'],
            ['📋 ESTADO DE LA SOLICITUD:', '✅ EN PROCESO' if entrevista.primer_nombre else '⚠️ PENDIENTE'],
        ]
        
        tabla_producto = Table(datos_producto, colWidths=[2.5*inch, 4.5*inch])
        tabla_producto.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_producto)
        story.append(Spacer(1, 20))
        
        # =================== SECCIÓN VII: AUTORIZACIONES Y CONSENTIMIENTOS ===================
        story.append(Paragraph("VII. AUTORIZACIONES Y CONSENTIMIENTOS", section_header_style))
        
        autoriza_apc = '✅ SÍ' if getattr(entrevista, 'autoriza_apc', None) else '❌ NO' if getattr(entrevista, 'autoriza_apc', None) is False else '☐ SÍ  ☐ NO'
        acepta_datos = '✅ SÍ' if getattr(entrevista, 'acepta_datos', None) else '❌ NO' if getattr(entrevista, 'acepta_datos', None) is False else '☐ SÍ  ☐ NO'
        
        datos_autorizaciones = [
            ['🔍 AUTORIZA CONSULTA APC:', autoriza_apc],
            ['🛡️ ACEPTA TRATAMIENTO DE DATOS:', acepta_datos],
            ['📋 AUTORIZA VERIFICACIONES:', '☐ SÍ  ☐ NO'],
            ['📞 ACEPTA CONTACTO TELEFÓNICO:', '☐ SÍ  ☐ NO'],
        ]
        
        tabla_autorizaciones = Table(datos_autorizaciones, colWidths=[2.5*inch, 4.5*inch])
        tabla_autorizaciones.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), pacifico_blue),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, pacifico_green),
            ('BACKGROUND', (0, 0), (0, -1), pacifico_light_blue),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, pacifico_light_green] * 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_autorizaciones)
        story.append(Spacer(1, 30))
        
        # =================== SECCIÓN DE FIRMAS PREMIUM ===================
        story.append(Paragraph("FIRMAS Y VALIDACIÓN OFICIAL", section_header_style))
        story.append(Spacer(1, 15))
        
        # Tabla de firmas con diseño premium
        firmas_data = [
            ['', '', ''],
            ['_________________________________', '', '_________________________________'],
            ['FIRMA DEL CLIENTE', '', 'FIRMA DEL OFICIAL DE CRÉDITO'],
            ['', '', ''],
            [f"{entrevista.primer_nombre} {entrevista.primer_apellido}".upper(), '', (entrevista.oficial or 'OFICIAL ASIGNADO').upper()],
            [f"C.I.P. {entrevista.provincia_cedula or '__'}-{entrevista.tipo_letra or '__'}-{entrevista.tomo_cedula or '____'}-{entrevista.partida_cedula or '_____'}", '', 'FINANCIERA PACÍFICO S.A.'],
        ]
        
        firma_table = Table(firmas_data, colWidths=[3*inch, 1*inch, 3*inch])
        firma_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 2), (0, 2), 'Helvetica-Bold'),
            ('FONTNAME', (2, 2), (2, 2), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 2), (0, 2), pacifico_blue),
            ('TEXTCOLOR', (2, 2), (2, 2), pacifico_blue),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 4), (-1, 4), 9),
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica'),
            ('FONTSIZE', (0, 5), (-1, 5), 8),
            ('TEXTCOLOR', (0, 5), (-1, 5), HexColor('#666666')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(firma_table)
        story.append(Spacer(1, 20))
        
        # Footer con fecha y validación
        footer_data = [
            [f"📅 Fecha: {fecha_actual.strftime('%d de %B de %Y')}", f"🕐 Hora: {fecha_actual.strftime('%H:%M hrs')}", f"📄 Doc: PDF-{entrevista_id:06d}"]
        ]
        
        footer_table = Table(footer_data, colWidths=[2.33*inch, 2.33*inch, 2.34*inch])
        footer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TEXTCOLOR', (0, 0), (-1, -1), pacifico_green),
            ('BACKGROUND', (0, 0), (-1, -1), pacifico_light_green),
            ('GRID', (0, 0), (-1, -1), 1, pacifico_green),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(footer_table)
        
        # Generar PDF
        doc.build(story)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"FormularioPremium_PAC{entrevista_id:06d}_{entrevista.primer_nombre}_{entrevista.primer_apellido}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print("🌟 PDF PREMIUM generado exitosamente - ¡El mejor del mundo! 🌟")
        return response
        
    except Exception as e:
        print(f"Error en descargar_entrevista_pdf: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)


def api_entrevistas(request):
    """API para obtener todas las entrevistas en formato JSON"""
    entrevistas = ClienteEntrevista.objects.all()
    campos = [field.name for field in ClienteEntrevista._meta.fields]
    data = []
    for entrevista in entrevistas:
        registro = {field: getattr(entrevista, field, None) for field in campos}
        data.append(registro)
    return JsonResponse({'entrevistas': data})
