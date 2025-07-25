from django.shortcuts import render, redirect, get_object_or_404
from django.forms import modelform_factory
from django.forms import inlineformset_factory
from .models import ClienteEntrevista, ReferenciaPersonal, ReferenciaComercial, OtroIngreso
from .forms import (
    ClienteEntrevistaForm,
    ReferenciaPersonalFormSet,
    ReferenciaComercialFormSet,
    OtroIngresoFormSet,
)
from django.http import HttpResponse, JsonResponse, Http404
from django.contrib import messages
from django.conf import settings
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import datetime
import csv
import json
import os
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO

# Main Views
def entrevista_cliente_view(request):
    """Test view to debug service worker issues"""
    import os
    from django.conf import settings
    
    sw_path = os.path.join(settings.BASE_DIR, 'workflow', 'static', 'workflow', 'sw.js')
    
    debug_info = {
        'sw_path': sw_path,
        'sw_exists': os.path.exists(sw_path),
        'base_dir': settings.BASE_DIR,
        'static_url': settings.STATIC_URL,
        'static_root': getattr(settings, 'STATIC_ROOT', 'Not set'),
        'debug': settings.DEBUG,
        'request_path': request.path,
        'request_full_path': request.get_full_path(),
        'request_host': request.get_host(),
    }
    
    if os.path.exists(sw_path):
        try:
            with open(sw_path, 'r', encoding='utf-8') as f:
                first_lines = f.read(500)  # Read first 500 chars
            debug_info['sw_content_preview'] = first_lines
            debug_info['sw_file_size'] = os.path.getsize(sw_path)
        except Exception as e:
            debug_info['sw_read_error'] = str(e)
    
    return JsonResponse(debug_info, json_dumps_params={'indent': 2})


def entrevista_cliente_view(request):
    steps_info = [
        ("Datos Generales", "fa-user"),
        ("Dirección", "fa-location-dot"),
        ("Cónyuge", "fa-people-roof"),
        ("Laboral", "fa-briefcase"),
        ("PEP", "fa-shield-halved"),
        ("Ref. Pers.", "fa-user-group"),
        ("Ref. Com.", "fa-building"),
        ("Bancarios", "fa-university"),
        ("Autorizaciones", "fa-circle-check"),
    ]


    if request.method == 'POST':
        form = ClienteEntrevistaForm(request.POST)
        referencias_formset = ReferenciaPersonalFormSet(
            request.POST,
            queryset=ReferenciaPersonal.objects.none(),
            prefix='referencias'
        )
        referencias_comerciales_formset = ReferenciaComercialFormSet(
            request.POST,
            queryset=ReferenciaComercial.objects.none(),
            prefix='referencias_comerciales'
        )
        otros_ingresos_formset = OtroIngresoFormSet(
            request.POST,
            queryset=OtroIngreso.objects.none(),
            prefix='otros_ingresos'
        )

        tiene_otros_ingresos = request.POST.get('tiene_otros_ingresos') == 'true'

        if (form.is_valid() and referencias_formset.is_valid() and
            referencias_comerciales_formset.is_valid() and otros_ingresos_formset.is_valid()):
            cliente = form.save(commit=False)
            
            # Asigna explícitamente todos los datos generales si no están en fields automáticos
            cliente.primer_nombre = form.cleaned_data.get('primer_nombre')
            cliente.segundo_nombre = form.cleaned_data.get('segundo_nombre')
            cliente.primer_apellido = form.cleaned_data.get('primer_apellido')
            cliente.segundo_apellido = form.cleaned_data.get('segundo_apellido')
            cliente.provincia_cedula = form.cleaned_data.get('provincia_cedula')
            cliente.tipo_letra = form.cleaned_data.get('tipo_letra')
            cliente.tomo_cedula = form.cleaned_data.get('tomo_cedula')
            cliente.partida_cedula = form.cleaned_data.get('partida_cedula')
            cliente.telefono = form.cleaned_data.get('telefono')
            cliente.email = form.cleaned_data.get('email')
            cliente.fecha_nacimiento = form.cleaned_data.get('fecha_nacimiento')
            cliente.sexo = form.cleaned_data.get('sexo')
            cliente.jubilado = form.cleaned_data.get('jubilado')
            cliente.nivel_academico = form.cleaned_data.get('nivel_academico')
            cliente.estado_civil = form.cleaned_data.get('estado_civil')
            cliente.no_dependientes = form.cleaned_data.get('no_dependientes')
            cliente.titulo = form.cleaned_data.get('titulo')
            cliente.salario = form.cleaned_data.get('salario')
            cliente.tipo_producto = form.cleaned_data.get('tipo_producto')
            cliente.oficial = form.cleaned_data.get('oficial')
            cliente.apellido_casada = form.cleaned_data.get('apellido_casada')
            cliente.peso = form.cleaned_data.get('peso')
            cliente.estatura = form.cleaned_data.get('estatura')
            cliente.nacionalidad = form.cleaned_data.get('nacionalidad')
            # ...otros campos generales si los tienes en el formulario...

            # Otros ingresos del modelo principal
            cliente.tipo_ingreso_1 = form.cleaned_data.get('tipo_ingreso_1')
            cliente.descripcion_ingreso_1 = form.cleaned_data.get('descripcion_ingreso_1')
            cliente.monto_ingreso_1 = form.cleaned_data.get('monto_ingreso_1')
            cliente.tipo_ingreso_2 = form.cleaned_data.get('tipo_ingreso_2')
            cliente.descripcion_ingreso_2 = form.cleaned_data.get('descripcion_ingreso_2')
            cliente.monto_ingreso_2 = form.cleaned_data.get('monto_ingreso_2')
            cliente.tipo_ingreso_3 = form.cleaned_data.get('tipo_ingreso_3')
            cliente.descripcion_ingreso_3 = form.cleaned_data.get('descripcion_ingreso_3')
            cliente.monto_ingreso_3 = form.cleaned_data.get('monto_ingreso_3')

            # Información laboral
            cliente.trabajo_direccion = form.cleaned_data.get('trabajo_direccion')
            cliente.trabajo_lugar = form.cleaned_data.get('trabajo_lugar')
            cliente.trabajo_cargo = form.cleaned_data.get('trabajo_cargo')
            cliente.salario = form.cleaned_data.get('salario')
            cliente.tipo_trabajo = form.cleaned_data.get('tipo_trabajo')
            cliente.frecuencia_pago = form.cleaned_data.get('frecuencia_pago')
            cliente.tel_trabajo = form.cleaned_data.get('tel_trabajo')
            cliente.tel_ext = form.cleaned_data.get('tel_ext')
            cliente.origen_fondos = form.cleaned_data.get('origen_fondos')
            cliente.fecha_inicio_trabajo = form.cleaned_data.get('fecha_inicio_trabajo')

            
            cliente.save()

            # Guardar referencias personales correctamente
            referencias = referencias_formset.save(commit=False)
            for ref in referencias:
                ref.entrevista = cliente
                ref.save()
            for ref in referencias_formset.deleted_objects:
                ref.delete()

            comerciales = referencias_comerciales_formset.save(commit=False)
            for com in comerciales:
                com.entrevista = cliente
                com.save()
            for com in referencias_comerciales_formset.deleted_objects:
                com.delete()

            if tiene_otros_ingresos:
                otros = otros_ingresos_formset.save(commit=False)
                for otro in otros:
                    otro.cliente = cliente
                    otro.save()
                for otro in otros_ingresos_formset.deleted_objects:
                    otro.delete()
            return redirect('formulario_gracias')
        else:
            # Print errors for debugging
            if form.errors:
                print('Errores en el formulario principal:', form.errors)
            print('Errores en referencias personales:', referencias_formset.errors)
            print('Errores en referencias comerciales:', referencias_comerciales_formset.errors)
            print('Errores en otros ingresos:', otros_ingresos_formset.errors)

    else:
        form = ClienteEntrevistaForm()
        referencias_formset = ReferenciaPersonalFormSet(
            queryset=ReferenciaPersonal.objects.none(),
            prefix='referencias'
        )
        referencias_comerciales_formset = ReferenciaComercialFormSet(
            queryset=ReferenciaComercial.objects.none(),
            prefix='referencias_comerciales'
        )
        otros_ingresos_formset = OtroIngresoFormSet(
            queryset=OtroIngreso.objects.none(),
            prefix='otros_ingresos'
        )

    return render(request, 'formulario/entrevista_cliente.html', {
        'form': form,
        'referencias_formset': referencias_formset,
        'referencias_comerciales_formset': referencias_comerciales_formset,
        'otros_ingresos_formset': otros_ingresos_formset,
        'steps_info': steps_info,
        'messages': messages.get_messages(request)
    })


def gracias(request):
    return render(request, 'formulario/gracias.html')


def lista_entrevistas(request):
    entrevistas = ClienteEntrevista.objects.values(
        'id', 'primer_nombre', 'primer_apellido', 'email', 'telefono', 'tipo_producto', 'oficial',
        'provincia_cedula', 'tipo_letra', 'tomo_cedula', 'partida_cedula', 'fecha_entrevista'
    )
    total_entrevistas = entrevistas.count()
    return render(request, 'formulario/lista_entrevistas.html', {
        'entrevistas': entrevistas,
        'total_entrevistas': total_entrevistas
    })


def descargar_entrevistas_excel(request):
    entrevistas = ClienteEntrevista.objects.all()

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    ws_entrevistas = wb.create_sheet(title="Entrevistas")
    campos_excluir = [
        'tipo_ingreso_1', 'descripcion_ingreso_1', 'monto_ingreso_1',
        'tipo_ingreso_2', 'descripcion_ingreso_2', 'monto_ingreso_2',
        'tipo_ingreso_3', 'descripcion_ingreso_3', 'monto_ingreso_3'
    ]
    campos_entrevista = [
        field for field in ClienteEntrevista._meta.fields
        if field.name != 'lugar_nacimiento' and field.name not in campos_excluir
    ]

    # --- AGREGADO: asegúrate de que peso y estatura estén en los encabezados ---
    encabezados = [
        field.verbose_name if field.name in ['peso', 'estatura'] else field.verbose_name.title() if field.name in ['conyuge_cargo', 'conyuge_ingreso'] else field.name
        for field in campos_entrevista
    ]
    if 'nacionalidad' not in [field.name for field in campos_entrevista]:
        encabezados.append('nacionalidad')
    # Si no están explícitamente, agrégalos al final
    if "Peso (lb)" not in encabezados:
        encabezados.append("Peso (lb)")
    if "Estatura (m)" not in encabezados:
        encabezados.append("Estatura (m)")
    ws_entrevistas.append(encabezados)

    for entrevista in entrevistas:
        fila = []
        for field in campos_entrevista:
            valor = getattr(entrevista, field.name, "")
            # Exporta peso y estatura como texto (sin conversión a decimal)
            if field.name in ['peso', 'estatura']:
                valor = str(valor) if valor is not None else ""
            # Para los toggles de autorizaciones y PEP, mostrar "Sí" o "No" en el Excel
            elif field.name in ['autoriza_apc', 'acepta_datos', 'es_beneficiario_final', 'es_pep', 'es_familiar_pep']:
                valor = "Sí" if valor else "No"
            elif field.name in ['conyuge_ingreso'] and valor is not None:
                try:
                    valor = float(valor)
                except Exception:
                    pass
            elif valor is not None and hasattr(valor, 'normalize'):
                valor = str(valor)
            if isinstance(valor, datetime.datetime) and valor.tzinfo is not None:
                valor = valor.replace(tzinfo=None)
            fila.append(valor if valor is not None else "")
        # Si no están explícitamente, agrégalos al final
        if 'nacionalidad' not in [field.name for field in campos_entrevista]:
            fila.append(getattr(entrevista, 'nacionalidad', ''))
        if 'peso' not in [field.name for field in campos_entrevista]:
            fila.append(str(getattr(entrevista, 'peso', '')) if getattr(entrevista, 'peso', None) is not None else "")
        if 'estatura' not in [field.name for field in campos_entrevista]:
            fila.append(str(getattr(entrevista, 'estatura', '')) if getattr(entrevista, 'estatura', None) is not None else "")
        ws_entrevistas.append(fila)

    ws_ref_personales = wb.create_sheet(title="Referencias Personales")
    campos_ref_personal = ['entrevista_id', 'nombre', 'telefono', 'relacion', 'direccion']
    ws_ref_personales.append(campos_ref_personal)

    for entrevista in entrevistas:
        for ref in entrevista.referencias_personales.all():
            fila = [entrevista.id, ref.nombre, ref.telefono, ref.relacion, ref.direccion]
            ws_ref_personales.append(fila)

    ws_ref_comerciales = wb.create_sheet(title="Referencias Comerciales")
    campos_ref_comercial = ['entrevista_id', 'tipo', 'nombre', 'actividad', 'telefono', 'celular', 'saldo']
    ws_ref_comerciales.append(campos_ref_comercial)

    for entrevista in entrevistas:
        for ref in entrevista.referencias_comerciales.all():
            fila = [entrevista.id, ref.tipo, ref.nombre, ref.actividad, ref.telefono, ref.celular, ref.saldo]
            ws_ref_comerciales.append(fila)

    ws_otros_ingresos = wb.create_sheet(title="Otros Ingresos")
    campos_otros_ingresos = ['cliente_id', 'tipo_ingreso', 'fuente', 'monto']
    ws_otros_ingresos.append(campos_otros_ingresos)

    for entrevista in entrevistas:
        for ingreso in entrevista.otros_ingresos.all():
            fila = [
                entrevista.id,
                ingreso.tipo_ingreso,
                ingreso.fuente,
                ingreso.monto
            ]
            ws_otros_ingresos.append(fila)

    for ws in [ws_entrevistas, ws_ref_personales, ws_ref_comerciales, ws_otros_ingresos]:
        for col_num, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=entrevistas_completas.xlsx'
    wb.save(response)
    return response


def descargar_entrevista_excel(request, entrevista_id):
    entrevista = ClienteEntrevista.objects.get(pk=entrevista_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="entrevista_{entrevista_id}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Campo', 'Valor'])

    for field in entrevista._meta.fields:
        writer.writerow([field.verbose_name, getattr(entrevista, field.name)])

    writer.writerow([])
    writer.writerow(['Referencias Personales'])
    for ref in entrevista.referencias_personales.all():
        writer.writerow([ref.nombre, ref.direccion, ref.relacion, '', '', '', ''])

    writer.writerow([])
    writer.writerow(['Referencias Comerciales'])
    for ref in entrevista.referencias_comerciales.all():
        writer.writerow([ref.nombre, ref.tipo, ref.actividad, ref.telefono, ref.celular, ref.saldo])

    writer.writerow([])
    writer.writerow(['Otros Ingresos'])
    for ingreso in entrevista.otros_ingresos.all():
        writer.writerow([ingreso.fuente, '', ingreso.monto])

    return response


def cliente_entrevista_create(request):
    if request.method == 'POST':
        form = ClienteEntrevistaForm(request.POST)
        otro_ingreso_formset = OtroIngresoFormSet(request.POST, prefix='otroingreso')
        referencia_personal_formset = ReferenciaPersonalFormSet(request.POST, prefix='refpersonal')
        referencia_comercial_formset = ReferenciaComercialFormSet(request.POST, prefix='refcomercial')
        if (form.is_valid() and otro_ingreso_formset.is_valid() and
                referencia_personal_formset.is_valid() and referencia_comercial_formset.is_valid()):
            cliente = form.save()
            otro_ingreso_formset.instance = cliente
            otro_ingreso_formset.save()
            referencia_personal_formset.instance = cliente
            referencia_personal_formset.save()
            referencia_comercial_formset.instance = cliente
            referencia_comercial_formset.save()
            return redirect('cliente_entrevista_detail', pk=cliente.pk)
    else:
        form = ClienteEntrevistaForm()
        otro_ingreso_formset = OtroIngresoFormSet(prefix='otroingreso')
        referencia_personal_formset = ReferenciaPersonalFormSet(prefix='refpersonal')
        referencia_comercial_formset = ReferenciaComercialFormSet(prefix='refcomercial')
    return render(request, 'workflow/cliente_entrevista_form.html', {
        'form': form,
        'otro_ingreso_formset': otro_ingreso_formset,
        'referencia_personal_formset': referencia_personal_formset,
        'referencia_comercial_formset': referencia_comercial_formset,
    })


def cliente_entrevista_detail(request, pk):
    cliente = get_object_or_404(ClienteEntrevista, pk=pk)
    return render(request, 'workflow/cliente_entrevista_detail.html', {'cliente': cliente})


def cliente_entrevista_list(request):
    clientes = ClienteEntrevista.objects.all()
    return render(request, 'workflow/cliente_entrevista_list.html', {'clientes': clientes})


def descargar_entrevista_pdf(request, entrevista_id):
    try:
        # Obtener la entrevista completa
        entrevista = get_object_or_404(ClienteEntrevista, id=entrevista_id)
        
        # Debug: verificar que tenemos la entrevista
        print(f"Entrevista encontrada: {entrevista.primer_nombre} {entrevista.primer_apellido}")
        
        # Crear buffer de memoria
        buffer = BytesIO()
        
        # Crear documento PDF simple para probar
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Agregar contenido básico
        story.append(Paragraph("FINANCIERA PACÍFICO", styles['Title']))
        story.append(Paragraph("FORMULARIO DE DATOS GENERALES", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        story.append(Paragraph(f"<b>Cliente:</b> {entrevista.primer_nombre} {entrevista.primer_apellido}", styles['Normal']))
        story.append(Spacer(1, 6))
        
        story.append(Paragraph(f"<b>Cédula:</b> {entrevista.provincia_cedula or ''}-{entrevista.tipo_letra or ''}-{entrevista.tomo_cedula or ''}-{entrevista.partida_cedula or ''}", styles['Normal']))
        story.append(Spacer(1, 6))
        
        story.append(Paragraph(f"<b>Email:</b> {entrevista.email or ''}", styles['Normal']))
        story.append(Spacer(1, 6))
        
        story.append(Paragraph(f"<b>Teléfono:</b> {entrevista.telefono or ''}", styles['Normal']))
        story.append(Spacer(1, 6))
        
        story.append(Paragraph(f"<b>Producto:</b> {entrevista.tipo_producto or ''}", styles['Normal']))
        story.append(Spacer(1, 6))
        
        story.append(Paragraph(f"<b>Oficial:</b> {entrevista.oficial or ''}", styles['Normal']))
        story.append(Spacer(1, 6))
        
        if entrevista.fecha_entrevista:
            story.append(Paragraph(f"<b>Fecha:</b> {entrevista.fecha_entrevista.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        
        # Información adicional disponible
        if hasattr(entrevista, 'salario') and entrevista.salario:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"<b>Salario:</b> B/. {entrevista.salario:,.2f}", styles['Normal']))
        
        if hasattr(entrevista, 'titulo') and entrevista.titulo:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"<b>Cargo:</b> {entrevista.titulo}", styles['Normal']))
        
        if hasattr(entrevista, 'nacionalidad') and entrevista.nacionalidad:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"<b>Nacionalidad:</b> {entrevista.nacionalidad}", styles['Normal']))
        
        # Espacio para firmas
        story.append(Spacer(1, 40))
        story.append(Paragraph("FIRMAS:", styles['Heading2']))
        story.append(Spacer(1, 30))
        
        # Tabla de firmas simple
        firmas_data = [
            ['_________________________', '', '_________________________'],
            ['FIRMA DEL CLIENTE', '', 'FIRMA DEL OFICIAL'],
            ['', '', ''],
            [f"{entrevista.primer_nombre} {entrevista.primer_apellido}", '', entrevista.oficial or ''],
            ['', '', ''],
            [f"Fecha: {datetime.datetime.now().strftime('%d/%m/%Y')}", '', '']
        ]
        
        firma_table = Table(firmas_data, colWidths=[2.5*inch, 1*inch, 2.5*inch])
        firma_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 1), (2, 1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        story.append(firma_table)
        
        # Generar PDF
        doc.build(story)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"formulario_{entrevista.primer_nombre}_{entrevista.primer_apellido}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print("PDF generado exitosamente")
        return response
        
    except Exception as e:
        print(f"Error en descargar_entrevista_pdf: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)


def api_entrevistas(request):
    entrevistas = ClienteEntrevista.objects.all()
    campos = [field.name for field in ClienteEntrevista._meta.fields]
    data = []
    for entrevista in entrevistas:
        registro = {field: getattr(entrevista, field, None) for field in campos}
        data.append(registro)
    return JsonResponse({'entrevistas': data})
