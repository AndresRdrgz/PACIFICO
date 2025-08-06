from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse, HttpResponseRedirect
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, F
from django.utils import timezone
import pytz
from django.contrib.auth.models import Group, User
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings
from django.db import models
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import io

from .modelsWorkflow import (
    Pipeline, Etapa, SubEstado, TransicionEtapa, PermisoEtapa, 
    Solicitud, HistorialSolicitud, Requisito, RequisitoPipeline, 
    RequisitoSolicitud, CampoPersonalizado, ValorCampoSolicitud,
    RequisitoTransicion, SolicitudComentario, PermisoPipeline, PermisoBandeja,
    NivelComite, UsuarioNivelComite, CatalogoPendienteAntesFirma, PendienteSolicitud,
    AgendaFirma
)
from .models import (
    HistorialBackoffice, CalificacionDocumentoBackoffice, 
    ComentarioDocumentoBackoffice, ClienteEntrevista, CalificacionCampo
)
from pacifico.models import UserProfile, Cliente, Cotizacion


def convertir_a_hora_panama(fecha_utc):
    """Convertir fecha UTC a hora de Panamá (UTC-5)"""
    if not fecha_utc:
        return 'N/A'
    
    try:
        # Zona horaria de Panamá (UTC-5)
        panama_tz = pytz.timezone('America/Panama')
        
        # Si la fecha ya tiene timezone info, convertir; si no, asumir UTC
        if fecha_utc.tzinfo is None:
            fecha_utc = pytz.utc.localize(fecha_utc)
        
        # Convertir a hora de Panamá
        fecha_panama = fecha_utc.astimezone(panama_tz)
        
        return fecha_panama.strftime('%d/%m/%Y %H:%M:%S')
    except Exception as e:
        print(f"Error convirtiendo fecha a Panamá: {e}")
        return 'Error'


def obtener_entradas_detalladas_backoffice():
    """
    Obtener cada ENTRADA a Back Office como una fila separada
    Incluye reprocesos como entradas independientes
    """
    entradas = []
    
    try:
        # Obtener todas las solicitudes que tienen historial de Back Office
        solicitudes_con_historial = HistorialBackoffice.objects.values_list('solicitud_id', flat=True).distinct()
        
        solicitudes_backoffice = Solicitud.objects.filter(
            id__in=solicitudes_con_historial
        ).select_related(
            'cotizacion', 'pipeline', 'etapa_actual', 'subestado_actual'
        )
        
        print(f"📊 Encontradas {solicitudes_backoffice.count()} solicitudes con historial Back Office")
        
        for solicitud in solicitudes_backoffice:
            # Obtener historial de Back Office ordenado por fecha
            historial_entries = HistorialBackoffice.objects.filter(
                solicitud=solicitud
            ).select_related('usuario', 'subestado_destino').order_by('fecha_entrada_subestado')
            
            print(f"📋 Solicitud {solicitud.codigo}: {historial_entries.count()} entradas de historial")
            
            if not historial_entries.exists():
                continue
            
            # Agrupar entradas por "ciclos" de Back Office
            # Un nuevo ciclo comienza cuando vuelve a Checklist después de salir
            ciclos = []
            ciclo_actual = []
            
            for entry in historial_entries:
                subestado_nombre = entry.subestado_destino.nombre if entry.subestado_destino else ''
                
                # Si llega a Checklist y ya hay un ciclo, iniciar nuevo ciclo
                if 'checklist' in subestado_nombre.lower() and ciclo_actual:
                    ciclos.append(ciclo_actual)
                    ciclo_actual = [entry]
                else:
                    ciclo_actual.append(entry)
            
            # Agregar el último ciclo
            if ciclo_actual:
                ciclos.append(ciclo_actual)
            
            print(f"🔄 Solicitud {solicitud.codigo}: {len(ciclos)} ciclos detectados")
            
            # Procesar cada ciclo como una entrada separada
            for idx, ciclo in enumerate(ciclos):
                entrada = procesar_ciclo_backoffice(solicitud, ciclo, idx + 1)
                if entrada:
                    entradas.append(entrada)
                    print(f"✅ Entrada {idx + 1} procesada para {solicitud.codigo}")
                    
    except Exception as e:
        print(f"Error obteniendo entradas detalladas: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"🎯 Total de entradas generadas: {len(entradas)}")
    return entradas


def procesar_ciclo_backoffice(solicitud, ciclo_historial, numero_entrada):
    """Procesar un ciclo completo en Back Office"""
    try:
        # Datos básicos de la solicitud
        entrada = {
            'codigo': solicitud.codigo,
            'cliente': solicitud.cliente_nombre_completo or (solicitud.cotizacion.nombreCliente if solicitud.cotizacion else 'N/A'),
            'cedula': solicitud.cotizacion.cedulaCliente if solicitud.cotizacion else 'N/A',
            'monto': f"B/. {solicitud.cotizacion.montoPrestamo:,.2f}" if solicitud.cotizacion and solicitud.cotizacion.montoPrestamo else 'N/A',
            'tipo_prestamo': solicitud.cotizacion.get_tipoPrestamo_display() if solicitud.cotizacion else 'N/A',
            'pipeline': solicitud.pipeline.nombre if solicitud.pipeline else 'N/A',
            'numero_entrada': numero_entrada,
            'es_reproceso': 'Sí' if numero_entrada > 1 else 'No',
            'motivo_devolucion': 'N/A'  # Se puede obtener de comentarios si está disponible
        }
        
        # Inicializar todos los campos de tiempo
        subestados = ['checklist', 'captura', 'firma', 'orden', 'tramite', 'subsanacion']
        for subestado in subestados:
            entrada[f'fecha_inicio_{subestado}'] = 'N/A'
            entrada[f'fecha_fin_{subestado}'] = 'N/A'
            entrada[f'usuario_{subestado}'] = 'N/A'
        
        entrada['fecha_llegada_bandeja'] = 'N/A'
        entrada['hora_asignacion'] = 'N/A'
        entrada['usuario_asignado'] = 'N/A'
        entrada['estado_final'] = 'N/A'
        entrada['tiempo_total'] = 'N/A'
        
        # Procesar cada entrada del historial en este ciclo
        fecha_inicio_ciclo = None
        fecha_fin_ciclo = None
        
        # Crear un mapa de subestados visitados en este ciclo
        subestados_visitados = {}
        
        for i, entry in enumerate(ciclo_historial):
            subestado_nombre = entry.subestado_destino.nombre.lower() if entry.subestado_destino else ''
            
            # Mapear nombres de subestados
            subestado_key = None
            if 'checklist' in subestado_nombre:
                subestado_key = 'checklist'
            elif 'captura' in subestado_nombre:
                subestado_key = 'captura'
            elif 'firma' in subestado_nombre:
                subestado_key = 'firma'
            elif 'orden' in subestado_nombre:
                subestado_key = 'orden'
            elif 'trámite' in subestado_nombre or 'tramite' in subestado_nombre:
                subestado_key = 'tramite'
            elif 'subsanación' in subestado_nombre or 'subsanacion' in subestado_nombre:
                subestado_key = 'subsanacion'
            
            if subestado_key:
                # Solo registrar la PRIMERA vez que entra a cada subestado en este ciclo
                if subestado_key not in subestados_visitados:
                    subestados_visitados[subestado_key] = {
                        'fecha_inicio': entry.fecha_entrada_subestado,
                        'usuario': entry.usuario,
                        'fecha_fin': entry.fecha_salida_subestado
                    }
                else:
                    # Si ya visitó este subestado, solo actualizar la fecha de salida si es más reciente
                    if entry.fecha_salida_subestado:
                        subestados_visitados[subestado_key]['fecha_fin'] = entry.fecha_salida_subestado
                
                # Primera entrada del ciclo = llegada a bandeja mixta
                if i == 0:
                    entrada['fecha_llegada_bandeja'] = convertir_a_hora_panama(entry.fecha_entrada_subestado)
                    fecha_inicio_ciclo = entry.fecha_entrada_subestado
                
                # Última salida del ciclo
                if entry.fecha_salida_subestado:
                    fecha_fin_ciclo = entry.fecha_salida_subestado
        
        # Llenar los datos finales basados en los subestados visitados
        for subestado_key, datos in subestados_visitados.items():
            entrada[f'fecha_inicio_{subestado_key}'] = convertir_a_hora_panama(datos['fecha_inicio'])
            entrada[f'usuario_{subestado_key}'] = datos['usuario'].get_full_name() if datos['usuario'] else 'N/A'
            if datos['fecha_fin']:
                entrada[f'fecha_fin_{subestado_key}'] = convertir_a_hora_panama(datos['fecha_fin'])
        
        # Calcular tiempo total del ciclo
        if fecha_inicio_ciclo and fecha_fin_ciclo:
            delta = fecha_fin_ciclo - fecha_inicio_ciclo
            horas_total = delta.total_seconds() / 3600
            entrada['tiempo_total'] = f"{horas_total:.2f} hrs"
        
        print(f"✅ Procesado ciclo {numero_entrada} para {solicitud.codigo}: {len(subestados_visitados)} subestados visitados")
        
        return entrada
        
    except Exception as e:
        print(f"Error procesando ciclo: {e}")
        import traceback
        traceback.print_exc()
        return None


@login_required
@csrf_exempt
def api_exportar_backoffice(request):
    """
    API para exportar estadísticas completas de Back Office en formato Excel
    con múltiples pestañas: Solicitudes y Tiempos, Documentos Problemáticos
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.db.models import Prefetch
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Eliminar la hoja por defecto
        wb.remove(wb.active)
        
        # === PESTAÑA 1: SOLICITUDES Y TIEMPOS ===
        ws_solicitudes = wb.create_sheet("Solicitudes y Tiempos")
        
        # Obtener todas las entradas detalladas a Back Office
        entradas_backoffice = obtener_entradas_detalladas_backoffice()
        
        # Encabezados para solicitudes (UNA FILA POR ENTRADA)
        headers_solicitudes = [
            'Código', 'Cliente', 'Cédula', 'Monto', 'Tipo Préstamo', 'Pipeline',
            'Número Entrada', 'Fecha Llegada Bandeja Mixta', 'Fecha Inicio Checklist', 'Fecha Fin Checklist', 'Usuario Checklist',
            'Fecha Inicio Captura', 'Fecha Fin Captura', 'Usuario Captura',
            'Fecha Inicio Firma', 'Fecha Fin Firma', 'Usuario Firma',
            'Fecha Inicio Orden', 'Fecha Fin Orden', 'Usuario Orden',
            'Fecha Inicio Trámite', 'Fecha Fin Trámite', 'Usuario Trámite',
            'Fecha Inicio Subsanación', 'Fecha Fin Subsanación', 'Usuario Subsanación',
            'Tiempo Total (hrs)', 'Es Reproceso', 'Motivo Devolución'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers_solicitudes, 1):
            cell = ws_solicitudes.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Procesar cada ENTRADA a Back Office (una fila por entrada)
        row = 2
        for entrada in entradas_backoffice:
            # Datos de la entrada
            datos = [
                entrada['codigo'],
                entrada['cliente'],
                entrada['cedula'],
                entrada['monto'],
                entrada['tipo_prestamo'],
                entrada['pipeline'],
                entrada['numero_entrada'],
                entrada['fecha_llegada_bandeja'],
                entrada['fecha_inicio_checklist'],
                entrada['fecha_fin_checklist'],
                entrada['usuario_checklist'],
                entrada['fecha_inicio_captura'],
                entrada['fecha_fin_captura'],
                entrada['usuario_captura'],
                entrada['fecha_inicio_firma'],
                entrada['fecha_fin_firma'],
                entrada['usuario_firma'],
                entrada['fecha_inicio_orden'],
                entrada['fecha_fin_orden'],
                entrada['usuario_orden'],
                entrada['fecha_inicio_tramite'],
                entrada['fecha_fin_tramite'],
                entrada['usuario_tramite'],
                entrada['fecha_inicio_subsanacion'],
                entrada['fecha_fin_subsanacion'],
                entrada['usuario_subsanacion'],
                entrada['tiempo_total'],
                entrada['es_reproceso'],
                entrada['motivo_devolucion']
            ]
            
            # Escribir datos en Excel
            for col, valor in enumerate(datos, 1):
                ws_solicitudes.cell(row=row, column=col, value=valor)
            
            row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers_solicitudes) + 1):
            ws_solicitudes.column_dimensions[get_column_letter(col)].width = 15
        
        # === PESTAÑA 2: DOCUMENTOS PROBLEMÁTICOS ===
        ws_documentos = wb.create_sheet("Documentos Problemáticos")
        
        # Encabezados para documentos
        headers_documentos = [
            'Código Solicitud', 'Cliente', 'Requisito/Documento', 'Estado Calificación',
            'Calificado Por', 'Fecha Calificación', 'Comentarios', 'Subestado Actual',
            'Tipo Problema', 'Días Pendiente'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers_documentos, 1):
            cell = ws_documentos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Obtener documentos problemáticos
        documentos_problematicos = obtener_documentos_problematicos()
        
        row = 2
        for doc in documentos_problematicos:
            datos_doc = [
                doc['codigo_solicitud'],
                doc['cliente'],
                doc['requisito'],
                doc['estado'],
                doc['calificado_por'],
                doc['fecha_calificacion'],
                doc['comentarios'],
                doc['subestado_actual'],
                doc['tipo_problema'],
                doc['dias_pendiente']
            ]
            
            for col, valor in enumerate(datos_doc, 1):
                cell = ws_documentos.cell(row=row, column=col, value=valor)
                # Colorear según el tipo de problema
                if doc['estado'] == 'malo':
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                elif doc['estado'] == 'pendiente':
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
            row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers_documentos) + 1):
            ws_documentos.column_dimensions[get_column_letter(col)].width = 18
        
        # === PESTAÑA 3: CHECKLIST PENDIENTES ===
        ws_checklist = wb.create_sheet("Checklist Pendientes")
        
        # Obtener elementos de checklist pendientes
        elementos_checklist = obtener_checklist_pendientes()
        
        headers_checklist = [
            'Código Solicitud', 'Cliente', 'Elemento Checklist', 'Estado',
            'Asignado A', 'Fecha Creación', 'Días Pendiente', 'Prioridad'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers_checklist, 1):
            cell = ws_checklist.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 2
        for elemento in elementos_checklist:
            datos_checklist = [
                elemento['codigo_solicitud'],
                elemento['cliente'],
                elemento['elemento'],
                elemento['estado'],
                elemento['asignado_a'],
                elemento['fecha_creacion'],
                elemento['dias_pendiente'],
                elemento['prioridad']
            ]
            
            for col, valor in enumerate(datos_checklist, 1):
                ws_checklist.cell(row=row, column=col, value=valor)
            
            row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers_checklist) + 1):
            ws_checklist.column_dimensions[get_column_letter(col)].width = 16
        
        # Preparar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="estadisticas_backoffice_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        import traceback
        print(f"Error en api_exportar_backoffice: {e}")
        print(traceback.format_exc())
        return JsonResponse({'error': f'Error al generar el Excel: {str(e)}'}, status=500)


def calcular_tiempos_subestados_backoffice(solicitud):
    """Calcular tiempos por subestado en Back Office usando HistorialBackoffice"""
    tiempos = {}
    
    try:
        # Obtener historial de Back Office para esta solicitud - tipo 'subestado'
        historial_entries = HistorialBackoffice.objects.filter(
            solicitud=solicitud,
            tipo_evento='subestado'
        ).select_related('subestado_destino', 'usuario').order_by('fecha_entrada_subestado')
        
        for entry in historial_entries:
            subestado_nombre = entry.subestado_destino.nombre if entry.subestado_destino else 'Sin Subestado'
            
            if entry.fecha_entrada_subestado and entry.fecha_salida_subestado:
                # Calcular tiempo transcurrido
                tiempo_transcurrido = entry.fecha_salida_subestado - entry.fecha_entrada_subestado
                horas = int(tiempo_transcurrido.total_seconds() // 3600)
            elif entry.fecha_entrada_subestado:
                # Si aún está en el subestado, calcular hasta ahora
                tiempo_transcurrido = timezone.now() - entry.fecha_entrada_subestado
                horas = int(tiempo_transcurrido.total_seconds() // 3600)
            else:
                horas = 0
            
            tiempos[subestado_nombre] = {
                'horas': horas,
                'usuario': entry.usuario.get_full_name() if entry.usuario else 'Sin Asignar',
                'fecha_entrada': entry.fecha_entrada_subestado.strftime('%Y-%m-%d %H:%M') if entry.fecha_entrada_subestado else '',
                'fecha_salida': entry.fecha_salida_subestado.strftime('%Y-%m-%d %H:%M') if entry.fecha_salida_subestado else 'En curso'
            }
    
    except Exception as e:
        print(f"Error calculando tiempos de subestados: {e}")
    
    return tiempos


def contar_veces_en_backoffice(solicitud):
    """Contar cuántas veces ha ingresado a Back Office"""
    try:
        return HistorialBackoffice.objects.filter(solicitud=solicitud).count()
    except:
        return 1


def calcular_estado_sla(solicitud):
    """Calcular el estado del SLA actual"""
    try:
        if not solicitud.etapa_actual or not solicitud.etapa_actual.sla:
            return 'Sin SLA'
        
        # Obtener tiempo transcurrido desde la última actualización
        if solicitud.fecha_ultima_actualizacion:
            tiempo_transcurrido = timezone.now() - solicitud.fecha_ultima_actualizacion
            horas_transcurridas = tiempo_transcurrido.total_seconds() / 3600
            
            sla_horas = solicitud.etapa_actual.sla.total_seconds() / 3600
            
            if horas_transcurridas > sla_horas:
                return 'Vencido'
            elif horas_transcurridas > sla_horas * 0.8:
                return 'Próximo a Vencer'
            else:
                return 'OK'
        
        return 'Sin Fecha'
    except:
        return 'Error'


def calcular_tiempo_bandeja_mixta(solicitud):
    """Calcular tiempo en bandeja mixta (sin asignar)"""
    try:
        # Buscar períodos donde asignada_a era None
        historial_entries = solicitud.historial.filter(
            usuario_responsable__isnull=True
        ).order_by('fecha_inicio')
        
        tiempo_total = 0
        for entry in historial_entries:
            if entry.fecha_inicio and entry.fecha_fin:
                tiempo_transcurrido = entry.fecha_fin - entry.fecha_inicio
                tiempo_total += tiempo_transcurrido.total_seconds() / 3600
        
        return int(tiempo_total)
    except:
        return 0


def obtener_documentos_problematicos():
    """Obtener documentos calificados como 'malo' o 'pendiente'"""
    documentos = []
    
    try:
        # Obtener calificaciones problemáticas
        calificaciones = CalificacionDocumentoBackoffice.objects.filter(
            Q(estado='malo') | Q(estado='pendiente')
        ).select_related(
            'requisito_solicitud__solicitud',
            'requisito_solicitud__requisito',
            'calificado_por',
            'opcion_desplegable'
        ).order_by('-fecha_calificacion')
        
        for cal in calificaciones:
            solicitud = cal.requisito_solicitud.solicitud
            
            # Calcular días pendiente
            if cal.fecha_calificacion:
                dias_pendiente = (timezone.now() - cal.fecha_calificacion).days
            else:
                dias_pendiente = 0
            
            # Obtener comentarios de la opción desplegable si existe
            comentarios = ''
            if cal.opcion_desplegable:
                comentarios = cal.opcion_desplegable.nombre
            
            documentos.append({
                'codigo_solicitud': solicitud.codigo,
                'cliente': solicitud.cliente_nombre_completo or (solicitud.cotizacion.nombreCliente if solicitud.cotizacion else ''),
                'requisito': cal.requisito_solicitud.requisito.nombre,
                'estado': cal.estado,
                'calificado_por': cal.calificado_por.get_full_name() if cal.calificado_por else '',
                'fecha_calificacion': cal.fecha_calificacion.strftime('%Y-%m-%d') if cal.fecha_calificacion else '',
                'comentarios': comentarios,
                'subestado_actual': solicitud.subestado_actual.nombre if solicitud.subestado_actual else '',
                'tipo_problema': 'Documento Malo' if cal.estado == 'malo' else 'Documento Pendiente',
                'dias_pendiente': dias_pendiente
            })
    
    except Exception as e:
        print(f"Error obteniendo documentos problemáticos: {e}")
    
    return documentos


def obtener_checklist_pendientes():
    """Obtener elementos de checklist pendientes"""
    elementos = []
    
    try:
        # Obtener pendientes de solicitudes en Back Office que no están completados
        pendientes = PendienteSolicitud.objects.filter(
            solicitud__etapa_actual__nombre__icontains='back office'
        ).exclude(
            estado='listo'  # Excluir los que ya están listos
        ).select_related(
            'solicitud', 'pendiente', 'agregado_por'
        ).order_by('-fecha_agregado')
        
        for pendiente in pendientes:
            # Calcular días pendiente
            dias_pendiente = (timezone.now() - pendiente.fecha_agregado).days if pendiente.fecha_agregado else 0
            
            elementos.append({
                'codigo_solicitud': pendiente.solicitud.codigo,
                'cliente': pendiente.solicitud.cliente_nombre_completo or (pendiente.solicitud.cotizacion.nombreCliente if pendiente.solicitud.cotizacion else ''),
                'elemento': pendiente.pendiente.descripcion,
                'estado': pendiente.get_estado_display(),
                'asignado_a': pendiente.agregado_por.get_full_name() if pendiente.agregado_por else 'Sin Asignar',
                'fecha_creacion': pendiente.fecha_agregado.strftime('%Y-%m-%d') if pendiente.fecha_agregado else '',
                'dias_pendiente': dias_pendiente,
                'prioridad': 'Alta' if dias_pendiente > 3 else 'Normal'
            })
    
    except Exception as e:
        print(f"Error obteniendo checklist pendientes: {e}")
    
    return elementos