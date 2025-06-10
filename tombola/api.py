import openpyxl
import os
from io import BytesIO
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.shortcuts import redirect, render
from django.contrib import messages
from django.core.mail import EmailMessage

from .models import Cliente, Boleto, Tombola, FormularioTombola, CargaMasiva
from tombola.views import enviar_boleto_por_correo, generate_boleto_pdf


def fetch_boletos_by_cedula(request):
    cedula = request.GET.get('cedula')
    try:
        cliente = Cliente.objects.get(cedulaCliente=cedula)
        boletos = Boleto.objects.filter(cliente=cliente)
        boletos_data = [
            {
                'id': boleto.id,
                'fecha_creacion': boleto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
                'tombola': boleto.tombola.id,
                'canalOrigen': boleto.canalOrigen,
            }
            for boleto in boletos
        ]
        return JsonResponse({'success': True, 'boletos': boletos_data})
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Cliente no encontrado.'})


def descargar_plantilla(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla Clientes"
    headers = ["Cedula", "Nombre", "Apellido", "Correo Electrónico", "Cantidad boletos"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_clientes.xlsx'
    workbook.save(response)
    return response


def carga_masiva(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'success': False, 'message': 'Por favor, suba un archivo válido.'})

        contador_boletos_creados = 0

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    cedula, nombre, apellido, correo, cantidad_boletos = row
                    if not cedula or not nombre or not apellido or not cantidad_boletos:
                        print(f"[Fila {row_num}] Datos incompletos. Saltando fila.")
                        continue

                    nombre_completo = f"{nombre} {apellido}"
                    cliente, _ = Cliente.objects.get_or_create(
                        cedulaCliente=cedula,
                        defaults={'nombreCliente': nombre_completo}
                    )

                    tombola = Tombola.objects.first()
                    if not tombola:
                        return JsonResponse({'success': False, 'message': 'No hay tómbolas registradas.'})

                    for _ in range(int(cantidad_boletos)):
                        boleto = Boleto.objects.create(
                            cliente=cliente,
                            tombola=tombola,
                            canalOrigen='Carga masiva'
                        )
                        contador_boletos_creados += 1

                        try:
                            FormularioTombola.objects.create(
                                nombre=nombre,
                                apellido=apellido,
                                cedulaCliente=cedula,
                                correo_electronico=correo or "",
                                celular="",
                                oficial="",
                                tombola=tombola,
                                boleto_asociado=boleto
                            )
                        except Exception as formulario_error:
                            print(f"[ERROR] No se pudo crear FormularioTombola para boleto {boleto.id}: {formulario_error}")

                        if correo:
                            try:
                                enviar_boleto_por_correo_desde_excel(boleto, correo, nombre, apellido)
                            except Exception as correo_error:
                                print(f"[ERROR] No se pudo enviar correo para el boleto {boleto.id}: {correo_error}")
                except Exception as fila_error:
                    print(f"[Fila {row_num}] Error inesperado: {fila_error}")
                    continue

            CargaMasiva.objects.create(
                archivo=file.name,
                cantidad_registros=contador_boletos_creados,
                usuario=request.user.username if request.user.is_authenticated else "Anónimo"
            )

            return JsonResponse({'success': True, 'message': 'Carga masiva completada y correos enviados.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Ocurrió un error al procesar el archivo: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Método no permitido.'})


def download_participantes_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Formulario Participantes"
    headers = [
        "Número de Boleto", "Nombre", "Apellido", "Cédula", "Celular", "Correo Electrónico",
        "Edad", "Fecha Nacimiento", "Sexo", "Sector", "Salario", "Producto Interesado",
        "Monto Solicitado", "Oficial", "Canal de Origen", "Tómbola", "Fecha de Creación"
    ]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    boletos = Boleto.objects.select_related('cliente', 'tombola').all().order_by('fecha_creacion')

    for row_num, boleto in enumerate(boletos, start=2):
        formulario = getattr(boleto, 'formulario_origen', None)
        cliente = boleto.cliente

        # Nombre y Apellido
        if formulario and formulario.nombre and formulario.apellido:
            nombre = formulario.nombre
            apellido = formulario.apellido
        else:
            nombre_parts = cliente.nombreCliente.split() if cliente.nombreCliente else []
            nombre = nombre_parts[0] if len(nombre_parts) > 0 else ""
            apellido = " ".join(nombre_parts[1:]) if len(nombre_parts) > 1 else ""

        # Datos comunes
        cedula = cliente.cedulaCliente or ""
        celular = getattr(formulario, 'celular', "") or ""
        correo = getattr(formulario, 'correo_electronico', "") or ""
        edad = getattr(formulario, 'edad', "") or ""

        # Fecha de nacimiento
        nacimiento = ""
        if formulario and formulario.fecha_nacimiento:
            try:
                nacimiento = formulario.fecha_nacimiento.strftime('%Y-%m-%d')
            except:
                nacimiento = ""

        sexo = getattr(formulario, 'sexo', "") or ""
        sector = getattr(formulario, 'sector', "") or ""
        salario = getattr(formulario, 'salario', "") or ""
        producto = getattr(formulario, 'producto_interesado', "") or ""

        # Monto solicitado
        monto = ""
        if formulario and formulario.dinero_a_solicitar:
            try:
                monto = float(formulario.dinero_a_solicitar)
            except (ValueError, TypeError):
                monto = ""

        oficial = getattr(formulario, 'oficial', "") or ""
        canal = boleto.canalOrigen or ""
        tombola = boleto.tombola.nombre if boleto.tombola else ""

        try:
            fecha = boleto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')
        except:
            fecha = ""

        # Escribir en hoja Excel
        fila = [
            f"{boleto.id:06d}", nombre, apellido, cedula, celular, correo, edad, nacimiento, sexo,
            sector, salario, producto, monto, oficial, canal, tombola, fecha
        ]
        for col_num, value in enumerate(fila, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Formulario_Participantes.xlsx'
    workbook.save(response)
    return response



def download_boletos_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Boletos"
    headers = ["ID", "Cliente", "Tómbola", "Fecha de Creación"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    boletos = Boleto.objects.all()
    for row_num, boleto in enumerate(boletos, start=2):
        sheet.cell(row=row_num, column=1, value=boleto.id)
        sheet.cell(row=row_num, column=2, value=str(boleto.cliente))
        sheet.cell(row=row_num, column=3, value=str(boleto.tombola))
        sheet.cell(row=row_num, column=4, value=boleto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=boletos.xlsx'
    workbook.save(response)
    return response


def listar_boletos(request):
    boletos = Boleto.objects.all()
    boletos_data = [
        {
            'id': boleto.id,
            'cliente': str(boleto.cliente),
            'tombola': str(boleto.tombola),
            'fecha_creacion': boleto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
            'canal_origen': boleto.canalOrigen,
        }
        for boleto in boletos
    ]
    return JsonResponse({'success': True, 'boletos': boletos_data})


def listar_formularios(request):
    formularios = FormularioTombola.objects.all()
    formularios_data = [
        {
            'id': formulario.id,
            'nombre': formulario.nombre,
            'apellido': formulario.apellido,
            'celular': formulario.celular,
            'correo_electronico': formulario.correo_electronico,
            'oficial': formulario.oficial,
            'tombola': str(formulario.tombola),
            'fecha_creacion': formulario.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
        }
        for formulario in formularios
    ]
    return JsonResponse({'success': True, 'formularios': formularios_data})


def enviar_boleto_por_correo_desde_excel(boleto, correo, nombre, apellido):
    try:
        template_path = os.path.join(settings.BASE_DIR, "tombola/templates/boleto_template.pdf")
        output_path = os.path.join(settings.BASE_DIR, f"tombola/temp/boleto_{boleto.id}.pdf")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        generate_boleto_pdf(template_path, output_path, boleto.id)

        subject = "Tu boleto de participación"
        body = (
            f"Hola {nombre} {apellido},\n\n"
            f"Adjunto encontrarás tu boleto de participación con el ID: {boleto.id:06d}, por tu préstamo desembolsado.\n\n"
            "Saludos,\nEquipo de FPACIFICO"
        )

        email = EmailMessage(
            subject=subject,
            body=body,
            from_email="tombola@fpacifico.com",
            to=[correo],
            cc=["", "jacastillo@fpacifico.com"],
        )

        with open(output_path, "rb") as pdf:
            email.attach(f"boleto_{boleto.id}.pdf", pdf.read(), "application/pdf")

        email.send()

        if not settings.DEBUG:
            os.remove(output_path)

        print(f"Correo enviado a {correo}")
    except Exception as e:
        print(f"[ERROR] No se pudo enviar correo para el boleto {boleto.id}: {e}")
