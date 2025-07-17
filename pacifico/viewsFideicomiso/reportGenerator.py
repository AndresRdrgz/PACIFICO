from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook
import os
import tempfile
from ..models import Cotizacion
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch


@login_required
def generate_report_excel(request, numero_cotizacion):
    """Generate Excel report"""
    try:
        cotizacion = get_object_or_404(Cotizacion, NumeroCotizacion=numero_cotizacion)
        
        # Path to the static Excel file
        excel_path = os.path.join(settings.BASE_DIR, 'static/insumos', 'consultaPrestAuto.xlsx')
        
        if not os.path.exists(excel_path):
            return HttpResponse("Excel template not found.", status=404)
        
        # Load the workbook and populate it
        workbook = load_workbook(excel_path)
        
        # Select the sheet with name "COTIZADOR PREST. AUTO"
        if "COTIZADOR PREST. AUTO" in workbook.sheetnames:
            sheet = workbook["COTIZADOR PREST. AUTO"]
        else:
            return HttpResponse("Sheet 'COTIZADOR PREST. AUTO' not found.", status=404)
        
        # Populate the Excel sheet
        _populate_excel_sheet(sheet, cotizacion)
        
        # Save the workbook to a temporary file
        temp_file = os.path.join(settings.BASE_DIR, 'static', f'temp_consulta_{numero_cotizacion}.xlsx')
        workbook.save(temp_file)
        
        # Serve the file as a response
        filename = f"Consulta - {numero_cotizacion} - {cotizacion.nombreCliente}.xlsx"
        with open(temp_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
        # Clean up
        try:
            os.unlink(temp_file)
        except:
            pass
            
        return response
        
    except Exception as e:
        error_message = str(e)
        print(f"Error generating Excel: {error_message}, User: {request.user.username}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required 
def generate_report_pdf(request, numero_cotizacion):
    """Generate PDF report from Excel template"""
    try:
        cotizacion = get_object_or_404(Cotizacion, NumeroCotizacion=numero_cotizacion)
        
        # First, create the populated Excel file using the same logic as generate_report_excel
        excel_path = os.path.join(settings.BASE_DIR, 'static/insumos', 'consultaPrestAuto.xlsx')
        
        if not os.path.exists(excel_path):
            return HttpResponse("Excel template not found.", status=404)
        
        # Load and populate the workbook
        workbook = load_workbook(excel_path)
        
        # Select the sheet with name "COTIZADOR PREST. AUTO"
        if "COTIZADOR PREST. AUTO" in workbook.sheetnames:
            sheet = workbook["COTIZADOR PREST. AUTO"]
        else:
            return HttpResponse("Sheet 'COTIZADOR PREST. AUTO' not found.", status=404)
        
        # Populate the Excel sheet with the same data as Excel export
        _populate_excel_sheet(sheet, cotizacion)
        
        # Create PDF from the populated Excel data
        pdf_path = os.path.join(settings.BASE_DIR, 'static', f'temp_consulta_{numero_cotizacion}.pdf')
        _create_pdf_from_excel_data(pdf_path, sheet, cotizacion, numero_cotizacion)
        
        # Return PDF as response
        with open(pdf_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            filename = f"Consulta - {numero_cotizacion} - {cotizacion.nombreCliente}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
        # Clean up
        try:
            os.unlink(pdf_path)
        except:
            pass
            
        return response
        
    except Exception as e:
        error_message = str(e)
        print(f"Error generating PDF: {error_message}, User: {request.user.username}")
        import traceback
        traceback.print_exc()  # This will help debug the actual error
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def _populate_excel_sheet(sheet, cotizacion):
    """Helper function to populate Excel sheet with cotizacion data"""
    try:
        # Populate basic fields
        if cotizacion.oficial:
            sheet['D6'] = cotizacion.oficial
        if cotizacion.nombreCliente:
            sheet['C10'] = cotizacion.nombreCliente
        if cotizacion.cedulaCliente:
            sheet['G10'] = cotizacion.cedulaCliente
        if cotizacion.tipoDocumento:
            sheet['H10'] = cotizacion.tipoDocumento
        if cotizacion.edad:
            sheet['J10'] = cotizacion.edad
        if cotizacion.sexo:
            sheet['I10'] = cotizacion.sexo
        if cotizacion.apcScore:
            sheet['k10'] = cotizacion.apcScore
        if cotizacion.apcPI:
            sheet['l10'] = cotizacion.apcPI / 100
        
        # Parameters
        if cotizacion.plazoPago:
            sheet['F14'] = cotizacion.plazoPago
        if cotizacion.r1:
            sheet['G14'] = cotizacion.r1 / 100
        if cotizacion.abonoPorcentaje:
            sheet['E14'] = cotizacion.abonoPorcentaje / 100
        if cotizacion.abono:
            sheet['e15'] = cotizacion.abono
        if cotizacion.tasaEstimada:
            sheet['H14'] = cotizacion.tasaEstimada / 100
        if cotizacion.cashback:
            sheet['E20'] = cotizacion.cashback
        if cotizacion.valorAuto:
            sheet['C14'] = cotizacion.valorAuto
        if cotizacion.calcMontoTimbres:
            sheet['L14'] = cotizacion.calcMontoTimbres
        
        sheet['i15'] = 'SI APLICA'
        if cotizacion.tasaBruta:
            sheet['J15'] = cotizacion.tasaBruta
            if cotizacion.tasaBruta == 0:
                sheet['K15'] = 'NO'
        
        # Details of calculation
        if cotizacion.montoPrestamo:
            sheet['E21'] = cotizacion.montoPrestamo
        if cotizacion.calcMontoNotaria:
            sheet['E23'] = cotizacion.calcMontoNotaria
        
        sheet['E24'] = 50  # promoPublicidad
        
        if cotizacion.mesesFinanciaSeguro and cotizacion.montoMensualSeguro:
            sheet['e26'] = cotizacion.mesesFinanciaSeguro * cotizacion.montoMensualSeguro
        if cotizacion.calcComiCierreFinal:
            sheet['e29'] = cotizacion.calcComiCierreFinal / 100
        if cotizacion.manejo_5porc:
            sheet['e30'] = cotizacion.manejo_5porc
        if cotizacion.auxMonto2:
            sheet['e31'] = cotizacion.auxMonto2
        if cotizacion.wrkLetraSinSeguros:
            sheet['E39'] = cotizacion.wrkLetraSinSeguros
            sheet['E43'] = cotizacion.wrkLetraSinSeguros
        if cotizacion.wrkLetraSeguro:
            sheet['e40'] = cotizacion.wrkLetraSeguro
        if cotizacion.wrkMontoLetra:
            sheet['E41'] = cotizacion.wrkMontoLetra
        if cotizacion.montoMensualSeguro:
            sheet['e42'] = cotizacion.montoMensualSeguro
        
        if cotizacion.wrkMontoLetra and cotizacion.montoMensualSeguro:
            sheet['E44'] = cotizacion.wrkMontoLetra + cotizacion.montoMensualSeguro
        if cotizacion.tablaTotalPagos:
            sheet['E46'] = cotizacion.tablaTotalPagos
        
        # Vehicle data
        if cotizacion.marca:
            sheet['j23'] = cotizacion.marca
        if cotizacion.modelo:
            sheet['j24'] = cotizacion.modelo
        if cotizacion.yearCarro:
            sheet['j25'] = cotizacion.yearCarro
        if cotizacion.montoMensualSeguro:
            sheet['j30'] = cotizacion.montoMensualSeguro
        if cotizacion.montoanualSeguro:
            sheet['j31'] = cotizacion.montoanualSeguro
        if cotizacion.transmisionAuto:
            sheet['j26'] = cotizacion.transmisionAuto
        if cotizacion.nuevoAuto:
            sheet['j27'] = cotizacion.nuevoAuto
        if cotizacion.kilometrajeAuto:
            sheet['j28'] = cotizacion.kilometrajeAuto
        
        # Vendor data
        if cotizacion.vendedor:
            sheet['j18'] = cotizacion.vendedor
        if cotizacion.vendedorComision:
            sheet['j20'] = cotizacion.vendedorComision
        
        # Debtor data
        if cotizacion.salarioBaseMensual:
            sheet['e77'] = cotizacion.salarioBaseMensual
        if cotizacion.tiempoServicio:
            sheet['E49'] = cotizacion.tiempoServicio
        if cotizacion.ingresos:
            sheet['J49'] = cotizacion.ingresos
        if cotizacion.nombreEmpresa:
            sheet['E50'] = cotizacion.nombreEmpresa 
        if cotizacion.referenciasAPC:
            sheet['J50'] = cotizacion.referenciasAPC
        if cotizacion.cartera:
            sheet['e51'] = cotizacion.cartera
        if cotizacion.licencia:
            sheet['J51'] = cotizacion.licencia
        if cotizacion.posicion:
            sheet['E52'] = cotizacion.posicion
        if cotizacion.perfilUniversitario:
            sheet['E53'] = cotizacion.perfilUniversitario
        
        # Comments
        if cotizacion.observaciones:
            sheet['I44'] = cotizacion.observaciones
        
    except Exception as e:
        print(f"Error populating Excel sheet: {e}")
        raise


def _create_pdf_from_excel_data(pdf_path, sheet, cotizacion, numero_cotizacion):
    """Create a comprehensive PDF from Excel data"""
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, 
                           topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Create content
    content = []
    
    # Title and header
    title = Paragraph(f"<b>CONSULTA DE PRÉSTAMO AUTOMOTRIZ</b>", styles['Title'])
    content.append(title)
    content.append(Spacer(1, 0.2*inch))
    
    subtitle = Paragraph(f"Cotización No. {numero_cotizacion}", styles['Heading2'])
    content.append(subtitle)
    content.append(Spacer(1, 0.2*inch))
    
    # Client information section
    client_title = Paragraph("<b>INFORMACIÓN DEL CLIENTE</b>", styles['Heading2'])
    content.append(client_title)
    content.append(Spacer(1, 0.1*inch))
    
    client_data = [
        ['Campo', 'Valor'],
        ['Oficial', str(cotizacion.oficial or '')],
        ['Cliente', str(cotizacion.nombreCliente or '')],
        ['Cédula', str(cotizacion.cedulaCliente or '')],
        ['Tipo Documento', str(cotizacion.tipoDocumento or '')],
        ['Edad', str(cotizacion.edad or '')],
        ['Sexo', str(cotizacion.sexo or '')],
        ['APC Score', str(cotizacion.apcScore or '')],
        ['Empresa', str(cotizacion.nombreEmpresa or '')],
        ['Tiempo de Servicio', str(cotizacion.tiempoServicio or '')],
        ['Salario Base', f"${cotizacion.salarioBaseMensual:,.2f}" if cotizacion.salarioBaseMensual else ''],
    ]
    
    client_table = Table(client_data, colWidths=[2.5*inch, 3*inch])
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    content.append(client_table)
    content.append(Spacer(1, 0.3*inch))
    
    # Vehicle information section
    vehicle_title = Paragraph("<b>INFORMACIÓN DEL VEHÍCULO</b>", styles['Heading2'])
    content.append(vehicle_title)
    content.append(Spacer(1, 0.1*inch))
    
    vehicle_data = [
        ['Campo', 'Valor'],
        ['Marca', str(cotizacion.marca or '')],
        ['Modelo', str(cotizacion.modelo or '')],
        ['Año', str(cotizacion.yearCarro or '')],
        ['Transmisión', str(cotizacion.transmisionAuto or '')],
        ['Condición', str(cotizacion.nuevoAuto or '')],
        ['Kilometraje', str(cotizacion.kilometrajeAuto or '')],
        ['Valor del Vehículo', f"${cotizacion.valorAuto:,.2f}" if cotizacion.valorAuto else ''],
    ]
    
    vehicle_table = Table(vehicle_data, colWidths=[2.5*inch, 3*inch])
    vehicle_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    content.append(vehicle_table)
    content.append(Spacer(1, 0.3*inch))
    
    # Main calculation results
    calc_title = Paragraph("<b>RESULTADOS DEL CÁLCULO</b>", styles['Heading2'])
    content.append(calc_title)
    content.append(Spacer(1, 0.1*inch))
    
    calculation_data = [
        ['Campo', 'Valor'],
        ['Monto del Préstamo', f"${cotizacion.montoPrestamo:,.2f}" if cotizacion.montoPrestamo else ''],
        ['Abono', f"${cotizacion.abono:,.2f}" if cotizacion.abono else ''],
        ['Plazo', f"{cotizacion.plazoPago} meses" if cotizacion.plazoPago else ''],
        ['Tasa de Interés', f"{cotizacion.tasaEstimada}%" if cotizacion.tasaEstimada else ''],
        ['Rentabilidad', f"{cotizacion.r1}%" if cotizacion.r1 else ''],
        ['Comisión de Cierre', f"{cotizacion.calcComiCierreFinal}%" if cotizacion.calcComiCierreFinal else ''],
        ['Letra Mensual (sin seguro)', f"${cotizacion.wrkLetraSinSeguros:,.2f}" if cotizacion.wrkLetraSinSeguros else ''],
        ['Colectivo de Crédito', f"${cotizacion.wrkLetraSeguro:,.2f}" if cotizacion.wrkLetraSeguro else ''],
        ['Letra APPX (con colectivo)', f"${cotizacion.wrkMontoLetra:,.2f}" if cotizacion.wrkMontoLetra else ''],
        ['Seguro de Auto (mensual)', f"${cotizacion.montoMensualSeguro:,.2f}" if cotizacion.montoMensualSeguro else ''],
        ['TOTAL LETRA CON SEGURO', f"${(cotizacion.wrkMontoLetra or 0) + (cotizacion.montoMensualSeguro or 0):,.2f}"],
        ['TOTAL A PAGAR', f"${cotizacion.tablaTotalPagos:,.2f}" if cotizacion.tablaTotalPagos else ''],
    ]
    
    calc_table = Table(calculation_data, colWidths=[3*inch, 2.5*inch])
    calc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -3), colors.lightblue),
        ('BACKGROUND', (0, -2), (-1, -1), colors.yellow),  # Highlight totals
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),  # Bold totals
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    content.append(calc_table)
    
    # Add observations if any
    if cotizacion.observaciones:
        content.append(Spacer(1, 0.2*inch))
        obs_title = Paragraph("<b>OBSERVACIONES</b>", styles['Heading3'])
        content.append(obs_title)
        content.append(Spacer(1, 0.1*inch))
        observations = Paragraph(str(cotizacion.observaciones), styles['Normal'])
        content.append(observations)
    
    # Build PDF
    doc.build(content)
