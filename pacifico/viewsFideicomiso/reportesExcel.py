from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook
from django.shortcuts import render, get_object_or_404
import os
from ..models import Cotizacion
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from fpdf import FPDF



@login_required
def generate_report(request, numero_cotizacion):
    # Retrieve the cotizacion record based on numero_cotizacion
    try:
        cotizacion = get_object_or_404(Cotizacion, NumeroCotizacion=numero_cotizacion)
        
        # Populate resultado with the values from the cotizacion record
        resultado = {
            'oficial': cotizacion.oficial,
            'nombreCliente': cotizacion.nombreCliente,
            'cedulaCliente': cotizacion.cedulaCliente,
            'tipoDocumento': cotizacion.tipoDocumento,
            'edad': cotizacion.edad,
            'sexo': cotizacion.sexo,
            'apcScore': cotizacion.apcScore,
            'apcPI': cotizacion.apcPI / 100 if cotizacion.apcPI is not None else '',
            'cotPlazoPago': cotizacion.plazoPago,
            'r1': cotizacion.r1,
            'abonoPorcentaje': cotizacion.abonoPorcentaje,
            'abono': cotizacion.abono,
            'cashback': cotizacion.cashback,
            'valorAuto': cotizacion.valorAuto,
            'calcMontoTimbres': cotizacion.calcMontoTimbres,
            'tasaBruta': cotizacion.tasaBruta,
            'aplicaPromocion': cotizacion.aplicaPromocion,
            'cotMontoPrestamo': cotizacion.montoPrestamo,
            'calcMontoNotaria': cotizacion.calcMontoNotaria,
            'promoPublicidad': 50,
            'montoLetraSeguroAdelantado': cotizacion.mesesFinanciaSeguro * cotizacion.montoMensualSeguro,
            'calcComiCierreFinal': cotizacion.calcComiCierreFinal,
            'manejo_5porc': cotizacion.manejo_5porc,
            'auxMonto2': cotizacion.auxMonto2,
            'wrkLetraSinSeguros': cotizacion.wrkLetraSinSeguros,
            'wrkLetraSeguro': cotizacion.wrkLetraSeguro,
            'wrkMontoLetra': cotizacion.wrkMontoLetra,
            'montoMensualSeguro': cotizacion.montoMensualSeguro,
            'wrkLetraConSeguros': cotizacion.wrkMontoLetra + cotizacion.montoMensualSeguro,
            'tablaTotalPagos': cotizacion.tablaTotalPagos,
            'vendedor': cotizacion.vendedor,
            'comisionVendedor': cotizacion.vendedorComision,
            'marcaAuto': cotizacion.marca,
            'lineaAuto': cotizacion.modelo,
            'yearAuto': cotizacion.yearCarro,
            'transmision': cotizacion.transmisionAuto,
            'nuevoAuto': cotizacion.nuevoAuto,
            'kilometrajeAuto': cotizacion.kilometrajeAuto,
            'observaciones': cotizacion.observaciones,
            'salarioBaseMensual': cotizacion.salarioBaseMensual,
            'tiempoServicio': cotizacion.tiempoServicio,
            'ingresos': cotizacion.ingresos,
            'nombreEmpresa': cotizacion.nombreEmpresa,
            'referenciasAPC': cotizacion.referenciasAPC,
            'cartera': cotizacion.cartera,
            'licencia': cotizacion.licencia,
            'posicion': cotizacion.posicion,
            'perfilUniversitario': cotizacion.perfilUniversitario,
            'horasExtrasMonto': cotizacion.horasExtrasMonto,
            'otrosMonto': cotizacion.otrosMonto,
            'montoanualSeguro': cotizacion.montoanualSeguro,
            'otrosDcto': cotizacion.otrosDcto,
            'bonosMonto': cotizacion.bonosMonto,
            'bonosDcto': cotizacion.bonosDcto,
            'siacapMonto': cotizacion.siacapMonto,
            'siacapDcto': cotizacion.siacapDcto,
            'praaMonto': cotizacion.praaMonto,
            'praaDcto': cotizacion.praaDcto,
            'dirOtrosMonto1': cotizacion.dirOtrosMonto1,
            'dirOtros1': cotizacion.dirOtros1,
            'dirOtrosDcto1': cotizacion.dirOtrosDcto1,
            'dirOtrosMonto2': cotizacion.dirOtrosMonto2,
            'dirOtros2': cotizacion.dirOtros2,
            'dirOtrosDcto2': cotizacion.dirOtrosDcto2,
            'dirOtrosMonto3': cotizacion.dirOtrosMonto3,
            'dirOtros3': cotizacion.dirOtros3,
            'dirOtrosDcto3': cotizacion.dirOtrosDcto3,
            'dirOtrosMonto4': cotizacion.dirOtrosMonto4,
            'dirOtros4': cotizacion.dirOtros4,
            'dirOtrosDcto4': cotizacion.dirOtrosDcto4,
            'pagoVoluntario1': cotizacion.pagoVoluntario1,
            'pagoVoluntarioMonto1': cotizacion.pagoVoluntarioMonto1,
            'pagoVoluntarioDcto1': cotizacion.pagoVoluntarioDcto1,
            'pagoVoluntario2': cotizacion.pagoVoluntario2,
            'pagoVoluntarioMonto2': cotizacion.pagoVoluntarioMonto2,
            'pagoVoluntarioDcto2': cotizacion.pagoVoluntarioDcto2,
            'pagoVoluntario3': cotizacion.pagoVoluntario3,
            'pagoVoluntarioMonto3': cotizacion.pagoVoluntarioMonto3,
            'pagoVoluntarioDcto3': cotizacion.pagoVoluntarioDcto3,
            'pagoVoluntario4': cotizacion.pagoVoluntario4,
            'pagoVoluntarioMonto4': cotizacion.pagoVoluntarioMonto4,
            'pagoVoluntarioDcto4': cotizacion.pagoVoluntarioDcto4,
            'pagoVoluntario5': cotizacion.pagoVoluntario5,
            'pagoVoluntarioMonto5': cotizacion.pagoVoluntarioMonto5,
            'pagoVoluntarioDcto5': cotizacion.pagoVoluntarioDcto5,
            'pagoVoluntario6': cotizacion.pagoVoluntario6,
            'pagoVoluntarioMonto6': cotizacion.pagoVoluntarioMonto6,
            'pagoVoluntarioDcto6': cotizacion.pagoVoluntarioDcto6,
            'mes0': cotizacion.mes0,
            'mes1': cotizacion.mes1,
            'mes2': cotizacion.mes2,
            'mes3': cotizacion.mes3,
            'mes4': cotizacion.mes4,
            'mes5': cotizacion.mes5,
            'mes6': cotizacion.mes6,
            'mes7': cotizacion.mes7,
            'mes8': cotizacion.mes8,
            'mes9': cotizacion.mes9,
            'mes10': cotizacion.mes10,
            'mes11': cotizacion.mes11,
            'primerMes': cotizacion.primerMes,
            'tipoProrrateo': cotizacion.tipoProrrateo,
            'tasaInteres': cotizacion.tasaEstimada / 100,
            'tablaTotalSeguro': cotizacion.tablaTotalSeguro,
            'tablaTotalInteres': cotizacion.tablaTotalInteres,
            'tablaTotalFeci': cotizacion.tablaTotalFeci,
            'numeroCotizacion': cotizacion.NumeroCotizacion,
            'montoManejoT': cotizacion.montoManejoT,
            'montoManejoB': cotizacion.monto_manejo_b,
            #codeudor
            'codeudorNombre': cotizacion.codeudorNombre,
            'codeudorCedula': cotizacion.codeudorCedula,
            'codeudorEstabilidad': cotizacion.codeudorEstabilidad,
            'codeudorIngresos': cotizacion.codeudorIngresos,
            'codeudorLicencia': cotizacion.codeudorLicencia,
            'codeudorEmpresa': cotizacion.codeudorEmpresa,
            'codeudorReferenciasAPC': cotizacion.codeudorReferenciasAPC,
            'codeudorNombreEmpres1': cotizacion.codeudorNombreEmpres1,
            'codeudorPeriodo1': cotizacion.codeudorPeriodo1,
            'codeudorSalario1': cotizacion.codeudorSalario1,
            'codeudorNombreEmpres2': cotizacion.codeudorNombreEmpres2,
            'codeudorPeriodo2': cotizacion.codeudorPeriodo2,
            'codeudorSalario2': cotizacion.codeudorSalario2,
            'codeudorNombreEmpres3': cotizacion.codeudorNombreEmpres3,
            'codeudorPeriodo3': cotizacion.codeudorPeriodo3,
            'codeudorSalario3': cotizacion.codeudorSalario3,
            'cobonosMonto': cotizacion.cobonosMonto,
            'cobonosDcto': cotizacion.cobonosDcto,
            'cohorasExtrasMonto': cotizacion.cohorasExtrasMonto,
            'cohorasExtrasDcto': cotizacion.cohorasExtrasDcto,
            'codeudorCartera': cotizacion.codeudorCartera,
            'codeudorPosicion': cotizacion.codeudorPosicion,
            'coprimaMonto': cotizacion.coprimaMonto,
            'coprimaDcto': cotizacion.coprimaDcto,
            'cootrosMonto': cotizacion.cootrosMonto,
            'cootrosDcto': cotizacion.cootrosDcto,
            'cosiacapMonto': cotizacion.cosiacapMonto,
            'cosiacapDcto': cotizacion.cosiacapDcto,
            'copraaMonto': cotizacion.copraaMonto,
            'copraaDcto': cotizacion.copraaDcto,
            'codirOtros1': cotizacion.codirOtros1,
            'codirOtrosMonto1': cotizacion.codirOtrosMonto1,
            'codirOtrosDcto1': cotizacion.codirOtrosDcto1,
            'codirOtros2': cotizacion.codirOtros2,
            'codirOtrosMonto2': cotizacion.codirOtrosMonto2,
            'codirOtrosDcto2': cotizacion.codirOtrosDcto2,
            'codirOtros3': cotizacion.codirOtros3,
            'codirOtrosMonto3': cotizacion.codirOtrosMonto3,
            'codirOtrosDcto3': cotizacion.codirOtrosDcto3,
            'codirOtros4': cotizacion.codirOtros4,
            'codirOtrosMonto4': cotizacion.codirOtrosMonto4,
            'codirOtrosDcto4': cotizacion.codirOtrosDcto4,
            'copagoVoluntario1': cotizacion.copagoVoluntario1,
            'copagoVoluntarioMonto1': cotizacion.copagoVoluntarioMonto1,
            'copagoVoluntarioDcto1': cotizacion.copagoVoluntarioDcto1,
            'copagoVoluntario2': cotizacion.copagoVoluntario2,
            'copagoVoluntarioMonto2': cotizacion.copagoVoluntarioMonto2,
            'copagoVoluntarioDcto2': cotizacion.copagoVoluntarioDcto2,
            'copagoVoluntario3': cotizacion.copagoVoluntario3,
            'copagoVoluntarioMonto3': cotizacion.copagoVoluntarioMonto3,
            'copagoVoluntarioDcto3': cotizacion.copagoVoluntarioDcto3,
            'copagoVoluntario4': cotizacion.copagoVoluntario4,
            'copagoVoluntarioMonto4': cotizacion.copagoVoluntarioMonto4,
            'copagoVoluntarioDcto4': cotizacion.copagoVoluntarioDcto4,
            'copagoVoluntario5': cotizacion.copagoVoluntario5,
            'copagoVoluntarioMonto5': cotizacion.copagoVoluntarioMonto5,
            'copagoVoluntarioDcto5': cotizacion.copagoVoluntarioDcto5,
            'copagoVoluntario6': cotizacion.copagoVoluntario6,
            'copagoVoluntarioMonto6': cotizacion.copagoVoluntarioMonto6,
            'copagoVoluntarioDcto6': cotizacion.copagoVoluntarioDcto6,



        }

        
        # Path to the static Excel file
        excel_path = os.path.join(settings.BASE_DIR, 'static/insumos', 'consultaPrestAuto.xlsx')

        if not os.path.exists(excel_path):
            return HttpResponse("File not found.", status=404)
        
        # Load the workbook and select the active sheet
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        # Select the sheet with name "COTIZADOR PREST. AUTO"
        if "COTIZADOR PREST. AUTO" in workbook.sheetnames:
            sheet = workbook["COTIZADOR PREST. AUTO"]
        else:
            return HttpResponse("Sheet not found.", status=404)
        
        # Example: Write the resultado data to the Excel sheet
        sheet['D6'] = resultado['oficial']
        sheet['C10'] = resultado['nombreCliente']
        sheet['G10'] = resultado['cedulaCliente']
        sheet['H10'] = resultado['tipoDocumento']
        sheet['J10'] = resultado['edad']
        sheet['I10'] = resultado['sexo']
        sheet['k10'] = resultado['apcScore']
        sheet['l10'] = resultado['apcPI']
        #parametros de la cotizacion
        sheet['F14'] = resultado['cotPlazoPago']
        sheet['G14'] = resultado['r1'] / 100
        sheet['E14'] = resultado['abonoPorcentaje'] / 100
        sheet['e15'] = resultado['abono']
        print('cashback', resultado['cashback'])
        sheet['H14'] = resultado['tasaInteres']
        sheet['E20'] = resultado['cashback']
        sheet['C14'] = resultado['valorAuto']
        sheet['L14'] = resultado['calcMontoTimbres']
        sheet['i15'] = 'SI APLICA'
        sheet['J15'] = resultado['tasaBruta']
        if resultado['tasaBruta'] == 0:
            sheet['K15'] = 'NO'

        #sheet['H14'] = resultado['cashback']

        #DETALLES DE LA COTIZACION
        sheet['E21'] = resultado['cotMontoPrestamo']
        sheet['E23'] = resultado['calcMontoNotaria']
        sheet['E24'] = resultado['promoPublicidad']
        sheet['e26'] = resultado['montoLetraSeguroAdelantado']
        sheet['e29'] = resultado['calcComiCierreFinal'] / 100
        sheet['e30'] = resultado['manejo_5porc']
        sheet['e31'] = resultado['auxMonto2']
        sheet['E39'] = resultado['wrkLetraSinSeguros']
        sheet['E43'] = resultado['wrkLetraSinSeguros']
        sheet['e40'] = resultado['wrkLetraSeguro']
        sheet['E41'] = resultado['wrkMontoLetra']
        sheet['e42'] = resultado['montoMensualSeguro']
        sheet['E44'] = resultado['wrkLetraConSeguros']

        sheet['E46'] = resultado['tablaTotalPagos']
        
        #DATOS DEL VENDEDOR
        sheet['j18'] = resultado['vendedor']
        sheet['j20'] = resultado['comisionVendedor']

        #DATOS DEL VEHICULO
        sheet['j23'] = resultado['marcaAuto']
        sheet['j24'] = resultado['lineaAuto']
        sheet['j25'] = resultado['yearAuto']
        sheet['j30'] = resultado['montoMensualSeguro']
        sheet['j31'] = resultado['montoanualSeguro']
        sheet['j26'] = resultado['transmision']
        sheet['j27'] = resultado['nuevoAuto']
        sheet['j28'] = resultado['kilometrajeAuto']

        #motivo consulta
        sheet['I44'] = resultado['observaciones']

        #promocion
        if resultado['aplicaPromocion'] == True:
             sheet['J42'] = 300
        else:
             sheet['J42'] = ''

        #DATOS DEL DEudor
        sheet['e77'] = resultado['salarioBaseMensual']
        sheet['E49']=resultado['tiempoServicio']
        sheet['J49']=resultado['ingresos']
        sheet['E50']=resultado['nombreEmpresa'] 
        sheet['J50']=resultado['referenciasAPC']
        sheet['e51']=resultado['cartera']
        sheet['J51']=resultado['licencia']
        sheet['E52']=resultado['posicion']
        sheet['E53']=resultado['perfilUniversitario']

        if resultado['horasExtrasMonto'] is None:
            resultado['horasExtrasMonto'] = 0
        if resultado['horasExtrasMonto'] > 0:
            sheet['J78']=resultado['horasExtrasMonto']

        if resultado['otrosMonto'] is None:
            resultado['otrosMonto'] = 0
        if resultado['otrosMonto'] > 0:
            sheet['J81']=resultado['otrosMonto']
            if resultado['otrosDcto'] == True:
                sheet['K81'] = 'SÍ'
            else:
                sheet['K81'] = 'NO'

        if resultado['bonosMonto'] is None:
            resultado['bonosMonto'] = 0
        if resultado['bonosMonto'] > 0:
            sheet['J80']=resultado['bonosMonto']
            if resultado['bonosDcto'] == True:
                sheet['K80'] = 'SÍ'
            else:
                sheet['K80'] = 'NO'
            

        

        #DESCUENTO DIRECTO
        
        print('siacapMonto', resultado['siacapMonto'])
        if resultado['siacapMonto'] is None:
            resultado['siacapMonto'] = 0
        if resultado['siacapMonto'] > 0:
            sheet['E87'] = resultado['siacapMonto']
            if resultado['siacapDcto'] == True:
                sheet['F87'] = 'SÍ'
            else:
                sheet['F87'] = 'NO'

        if resultado['praaMonto'] is None:
            resultado['praaMonto'] = 0
        if resultado['praaMonto'] > 0:
            sheet['E88'] = resultado['praaMonto']
            if resultado['praaDcto'] == True:
                sheet['F88'] = 'SÍ'
            else:
                sheet['F88'] = 'NO'
        
        if resultado['dirOtrosMonto1'] is None:
            resultado['dirOtrosMonto1'] = 0
        if resultado['dirOtrosMonto1'] > 0:
            sheet['E89'] = resultado['dirOtrosMonto1']
            sheet['C89'] = resultado['dirOtros1']
            if resultado['dirOtrosDcto1'] == True:
                sheet['F89'] = 'SÍ'
            else:
                sheet['F89'] = 'NO'

        if resultado['dirOtrosMonto2'] is None:
            resultado['dirOtrosMonto2'] = 0

        if resultado['dirOtrosMonto2'] > 0:
            sheet['E90'] = resultado['dirOtrosMonto2']
            sheet['C90'] = resultado['dirOtros2']
            if resultado['dirOtrosDcto2'] == True:
                sheet['F90'] = 'SÍ'
            else:
                sheet['F90'] = 'NO'

        if resultado['dirOtrosMonto3'] is None:
            resultado['dirOtrosMonto3'] = 0

        if resultado['dirOtrosMonto3'] > 0:
            sheet['E91'] = resultado['dirOtrosMonto3']
            sheet['C91'] = resultado['dirOtros3']
            if resultado['dirOtrosDcto3'] == True:
                sheet['F91'] = 'SÍ'
            else:
                sheet['F91'] = 'NO'

        if resultado['dirOtrosMonto4'] is None:
            resultado['dirOtrosMonto4'] = 0


        if resultado['dirOtrosMonto4'] > 0:
            sheet['E92'] = resultado['dirOtrosMonto4']
            sheet['C92'] = resultado['dirOtros4']
            if resultado['dirOtrosDcto4'] == True:
                sheet['F92'] = 'SÍ'
            else:
                sheet['F92'] = 'NO'

        #pagos voluntarios
        sheet['H87'] = resultado['pagoVoluntario1']
        sheet['H88'] = resultado['pagoVoluntario2']
        sheet['H89'] = resultado['pagoVoluntario3']
        
        sheet['H90'] = resultado['pagoVoluntario4']
        
        sheet['H91'] = resultado['pagoVoluntario5']
        
        sheet['H92'] = resultado['pagoVoluntario6']
        
        

        #PARSE TRUE TO SI AND FALSE TO NO
        if resultado ['pagoVoluntarioMonto1'] is None:
            resultado['pagoVoluntarioMonto1'] = 0

        if resultado['pagoVoluntarioMonto1'] > 0:
            sheet['J87'] = resultado['pagoVoluntarioMonto1']
            if resultado['pagoVoluntarioDcto1'] == True:
                sheet['K87'] = 'SÍ'
            else:
                sheet['K87'] = 'NO'

        if resultado['pagoVoluntarioMonto2'] is None:
            resultado['pagoVoluntarioMonto2'] = 0
        if resultado['pagoVoluntarioMonto2'] > 0:
            sheet['J88'] = resultado['pagoVoluntarioMonto2']
            if resultado['pagoVoluntarioDcto2'] == True:
                sheet['K88'] = 'SÍ'
            else:
                sheet['K88'] = 'NO'

        if resultado['pagoVoluntarioMonto3'] is None:
            resultado['pagoVoluntarioMonto3'] = 0
        if resultado['pagoVoluntarioMonto3'] > 0:
            sheet['J89'] = resultado['pagoVoluntarioMonto3']
            if resultado['pagoVoluntarioDcto3'] == True:
                sheet['K89'] = 'SÍ'
            else:
                sheet['K89'] = 'NO'
        if resultado['pagoVoluntarioMonto4'] is None:
            resultado['pagoVoluntarioMonto4'] = 0

        if resultado['pagoVoluntarioMonto4'] > 0:
            sheet['J90'] = resultado['pagoVoluntarioMonto4']
            if resultado['pagoVoluntarioDcto4'] == True:
                sheet['K90'] = 'SÍ'
            else:
                sheet['K90'] = 'NO'

        if resultado['pagoVoluntarioMonto5'] is None:
            resultado['pagoVoluntarioMonto5'] = 0

        if resultado['pagoVoluntarioMonto5'] > 0:
            sheet['J91'] = resultado['pagoVoluntarioMonto5']
            if resultado['pagoVoluntarioDcto5'] == True:
                sheet['K91'] = 'SÍ'
            else:
                sheet['K91'] = 'NO'
        
        if resultado['pagoVoluntarioMonto6'] is None:
            resultado['pagoVoluntarioMonto6'] = 0

        if resultado['pagoVoluntarioMonto6'] > 0:
            sheet['J92'] = resultado['pagoVoluntarioMonto6']
            if resultado['pagoVoluntarioDcto6'] == True:
                sheet['K92'] = 'SÍ'
            else:
                sheet['K92'] = 'NO'

        #Codeudor
        sheet['E103'] = resultado['codeudorNombre']
        sheet['E105'] = resultado['codeudorCedula']
        sheet['J103'] = resultado['codeudorCartera']
        sheet['J107'] = resultado['codeudorPosicion']

        sheet['E56'] = resultado['codeudorEstabilidad']
        sheet['E58'] = resultado['codeudorIngresos']
        sheet['J57'] = resultado['codeudorLicencia']
        sheet['E57'] = resultado['codeudorEmpresa']
        sheet['J56'] = resultado['codeudorReferenciasAPC']

        sheet['E61'] = resultado['codeudorNombreEmpres1']
        sheet['I61'] = resultado['codeudorPeriodo1']
        sheet['K61'] = resultado['codeudorSalario1']
        sheet['E62'] = resultado['codeudorNombreEmpres2']
        sheet['I62'] = resultado['codeudorPeriodo2']
        sheet['K62'] = resultado['codeudorSalario2']
        sheet['E63'] = resultado['codeudorNombreEmpres3']
        sheet['I63'] = resultado['codeudorPeriodo3']
        sheet['K63'] = resultado['codeudorSalario3']

        #CODEUDOR - INGRESOS ADICIONALES
        if resultado['cohorasExtrasMonto'] is None:
            resultado['cohorasExtrasMonto'] = 0
        if resultado['cohorasExtrasMonto'] > 0:
            sheet['J114']=resultado['cohorasExtrasMonto']
            if resultado['cohorasExtrasDcto'] == True:
                sheet['K114'] = 'SÍ'
            else:
                sheet['K114'] = 'NO'

        if resultado['coprimaMonto'] is None:
            resultado['coprimaMonto'] = 0
        if resultado['coprimaMonto'] > 0:
            sheet['J115']=resultado['coprimaMonto']
            if resultado['coprimaDcto'] == True:
                sheet['K115'] = 'SÍ'
            else:
                sheet['K115'] = 'NO'

        if resultado['cobonosMonto'] is None:
            resultado['cobonosMonto'] = 0
        if resultado['cobonosMonto'] > 0:
            sheet['J116']=resultado['cobonosMonto']
            if resultado['cobonosDcto'] == True:
                sheet['K116'] = 'SÍ'
            else:
                sheet['K116'] = 'NO'

        if resultado['cootrosMonto'] is None:
            resultado['cootrosMonto'] = 0
        if resultado['cootrosMonto'] > 0:
            sheet['J117']=resultado['cootrosMonto']
            if resultado['cootrosDcto'] == True:
                sheet['K117'] = 'SÍ'
            else:
                sheet['K117'] = 'NO'
    #codeudor - DEscuentos directos
        if resultado['cosiacapMonto'] is None:
                resultado['cosiacapMonto'] = 0
        if resultado['cosiacapMonto'] > 0:
                sheet['E123']=resultado['cosiacapMonto']
                if resultado['cosiacapDcto'] == True:
                    sheet['F123'] = 'SÍ'
                else:
                    sheet['F123'] = 'NO'

        if resultado['copraaMonto'] is None:
                resultado['copraaMonto'] = 0
        if resultado['copraaMonto'] > 0:
                sheet['E124']=resultado['copraaMonto']
                if resultado['copraaDcto'] == True:
                    sheet['F124'] = 'SÍ'
                else:
                    sheet['F124'] = 'NO'

        if resultado['codirOtrosMonto1'] is None:
                resultado['codirOtrosMonto1'] = 0
        if resultado['codirOtrosMonto1'] > 0:
                sheet['E125']=resultado['codirOtrosMonto1']
                sheet['C125']=resultado['codirOtros1']
                if resultado['codirOtrosDcto1'] == True:
                    sheet['F125'] = 'SÍ'
                else:
                    sheet['F125'] = 'NO'
        
        if resultado['codirOtrosMonto2'] is None:
                resultado['codirOtrosMonto2'] = 0
        if resultado['codirOtrosMonto2'] > 0:
                sheet['E126']=resultado['codirOtrosMonto2']
                sheet['C126']=resultado['codirOtros2']
                if resultado['codirOtrosDcto2'] == True:
                    sheet['F126'] = 'SÍ'
                else:
                    sheet['F126'] = 'NO'

        if resultado['codirOtrosMonto3'] is None:
                resultado['codirOtrosMonto3'] = 0
        if resultado['codirOtrosMonto3'] > 0:
                sheet['E127']=resultado['codirOtrosMonto3']
                sheet['C127']=resultado['codirOtros3']
                if resultado['codirOtrosDcto3'] == True:
                    sheet['F127'] = 'SÍ'
                else:
                    sheet['F127'] = 'NO'

        if resultado['codirOtrosMonto4'] is None:
                resultado['codirOtrosMonto4'] = 0
        if resultado['codirOtrosMonto4'] > 0:
                sheet['E128']=resultado['codirOtrosMonto4']
                sheet['C128']=resultado['codirOtros4']
                if resultado['codirOtrosDcto4'] == True:
                    sheet['F128'] = 'SÍ'
                else:
                    sheet['F128'] = 'NO'
    #CODEUDOR - PAGOS VOLUNTARIOS
        if resultado['copagoVoluntarioMonto1'] is None:
                resultado['copagoVoluntarioMonto1'] = 0
        if resultado['copagoVoluntarioMonto1'] > 0:
                sheet['J123']=resultado['copagoVoluntarioMonto1']
                sheet['H123']=resultado['copagoVoluntario1']
                if resultado['copagoVoluntarioDcto1'] == True:
                    sheet['K123'] = 'SÍ'
                else:
                    sheet['K123'] = 'NO'
        
        if resultado['copagoVoluntarioMonto2'] is None:
                resultado['copagoVoluntarioMonto2'] = 0
        if resultado['copagoVoluntarioMonto2'] > 0:
                sheet['J124']=resultado['copagoVoluntarioMonto2']
                sheet['H124']=resultado['copagoVoluntario2']
                if resultado['copagoVoluntarioDcto2'] == True:
                    sheet['K124'] = 'SÍ'
                else:
                    sheet['K124'] = 'NO'
        
        if resultado['copagoVoluntarioMonto3'] is None:
                resultado['copagoVoluntarioMonto3'] = 0
        if resultado['copagoVoluntarioMonto3'] > 0:
                sheet['J125']=resultado['copagoVoluntarioMonto3']
                sheet['H125']=resultado['copagoVoluntario3']
                if resultado['copagoVoluntarioDcto3'] == True:
                    sheet['K125'] = 'SÍ'
                else:
                    sheet['K125'] = 'NO'

        if resultado['copagoVoluntarioMonto4'] is None:
                resultado['copagoVoluntarioMonto4'] = 0
        if resultado['copagoVoluntarioMonto4'] > 0:
                sheet['J126']=resultado['copagoVoluntarioMonto4']
                sheet['H126']=resultado['copagoVoluntario4']
                if resultado['copagoVoluntarioDcto4'] == True:
                    sheet['K126'] = 'SÍ'
                else:
                    sheet['K126'] = 'NO'
        
        if resultado['copagoVoluntarioMonto5'] is None:
                resultado['copagoVoluntarioMonto5'] = 0
        if resultado['copagoVoluntarioMonto5'] > 0: 
                sheet['J127']=resultado['copagoVoluntarioMonto5']
                sheet['H127']=resultado['copagoVoluntario5']
                if resultado['copagoVoluntarioDcto5'] == True:
                    sheet['K127'] = 'SÍ'
                else:
                    sheet['K127'] = 'NO'

        if resultado['copagoVoluntarioMonto6'] is None: 
                resultado['copagoVoluntarioMonto6'] = 0
        if resultado['copagoVoluntarioMonto6'] > 0:
                sheet['J128']=resultado['copagoVoluntarioMonto6']
                sheet['H128']=resultado['copagoVoluntario6']
                if resultado['copagoVoluntarioDcto6'] == True:
                    sheet['K128'] = 'SÍ'
                else:
                    sheet['K128'] = 'NO'


     # Select the sheet with name "PRORRATEO"
        if "PRORRATEO" in workbook.sheetnames:
            prorrateo = workbook["PRORRATEO"]
        else:
            return HttpResponse("Sheet not found.", status=404)
        
        prorrateo['D6'] = resultado['mes0']
        prorrateo['d7'] = resultado['mes1']
        prorrateo['d8'] = resultado['mes2']
        prorrateo['d9'] = resultado['mes3']
        prorrateo['d10'] = resultado['mes4']
        prorrateo['d11'] = resultado['mes5']
        prorrateo['d12'] = resultado['mes6']
        prorrateo['d13'] = resultado['mes7']
        prorrateo['d14'] = resultado['mes8']
        prorrateo['d15'] = resultado['mes9']
        prorrateo['d16'] = resultado['mes10']
        prorrateo['d17'] = resultado['mes11']
        prorrateo['C6'] = resultado['primerMes']
        print('prmer mes', resultado['primerMes'])

         # Select the sheet with name "DESGLOSE"
        if "DESGLOSE" in workbook.sheetnames:
            desglose = workbook["DESGLOSE"]
        else:
            return HttpResponse("Sheet not found.", status=404)

        print ('Desglose sra Raquel')
        desglose['I4'] = resultado['numeroCotizacion']
        desglose['C6'] = resultado['r1'] / 100
        desglose['C7'] = resultado['tasaInteres']
        desglose['I8'] = str(resultado['cotPlazoPago']) + '/' + str(resultado['cotPlazoPago'])
        
        desglose['D12'] = resultado['auxMonto2']
        desglose['G16'] = resultado['manejo_5porc']

        desglose['C14'] = resultado['calcMontoNotaria']
        desglose['C15'] = resultado['cotMontoPrestamo']
        
        desglose['H12'] = resultado ['calcComiCierreFinal'] / 100
        desglose['C13'] = resultado['montoManejoT']
        desglose['G12'] = resultado['montoManejoT']
        desglose['G15'] = resultado['montoManejoB']
        desglose['G13'] = resultado['calcMontoTimbres']

        desglose['D18'] = resultado['tablaTotalSeguro']
        desglose['d19'] = resultado['tablaTotalInteres']
        desglose['d20'] = resultado['tablaTotalFeci']
        desglose['d21'] = resultado['tablaTotalPagos']

        # Select the sheet with name "mov bancarios"
        if "MOV. BANCARIOS" in workbook.sheetnames:
            movimientos = workbook["MOV. BANCARIOS"]
        else:
            return HttpResponse("Sheet not found.", status=404)
        
        movimientos['C6'] = cotizacion.movPrimerMes if cotizacion.movPrimerMes is not None else "ENERO"
        movimientos['C8'] = cotizacion.ingresosMes1 if cotizacion.ingresosMes1 is not None else 0
        movimientos['D8'] = cotizacion.egresosMes1 if cotizacion.egresosMes1 is not None else 0
        movimientos['f8'] = cotizacion.ingresosMes2 if cotizacion.ingresosMes2 is not None else 0
        movimientos['g8'] = cotizacion.egresosMes2 if cotizacion.egresosMes2 is not None else 0
        movimientos['i8'] = cotizacion.ingresosMes3 if cotizacion.ingresosMes3 is not None else 0
        movimientos['j8'] = cotizacion.egresosMes3 if cotizacion.egresosMes3 is not None else 0
        movimientos['l8'] = cotizacion.ingresosMes4 if cotizacion.ingresosMes4 is not None else 0
        movimientos['m8'] = cotizacion.egresosMes4 if cotizacion.egresosMes4 is not None else 0
        movimientos['o8'] = cotizacion.ingresosMes5 if cotizacion.ingresosMes5 is not None else 0
        movimientos['p8'] = cotizacion.egresosMes5 if cotizacion.egresosMes5 is not None else 0
        movimientos['r8'] = cotizacion.ingresosMes6 if cotizacion.ingresosMes6 is not None else 0
        movimientos['s8'] = cotizacion.egresosMes6 if cotizacion.egresosMes6 is not None else 0
        try:
            if cotizacion.movOpcion == "tomar_valor":
                movimientos['U13'] = "TOMAR VALOR"
            else:
                movimientos['U13'] = cotizacion.movOpcion if cotizacion.movOpcion is not None else 'COLOCAR MANUAL'
        except:
                movimientos['U13'] = "COLOCAR MANUAL"

        
        # Save the workbook to a temporary file
        temp_file = os.path.join(settings.BASE_DIR, 'static', 'temp_consultaFideicomiso.xlsx')
        workbook.save(temp_file)
        
        # Serve the file as a response
        nombre_cliente = resultado['nombreCliente']
        filename = f"Consulta - {numero_cotizacion} -{nombre_cliente}.xlsx"
        with open(temp_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response
    
    except Exception as e:
        error_message = str(e)
        # Log the error
        print(f"Error: {error_message}, User: {request.user.username}")
        return JsonResponse({'status': 'error', 'message': str(e), 'resultado': resultado}, status=500)
  


def generate_report_temp(numero_cotizacion):
    # Retrieve the cotizacion record based on numero_cotizacion
    cotizacion = get_object_or_404(Cotizacion, NumeroCotizacion=numero_cotizacion) 
    excel_path = os.path.join(settings.BASE_DIR, 'static/insumos', 'consultaPrestAuto.xlsx')
    if not os.path.exists(excel_path):
        raise FileNotFoundError("File not found.")
    
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    # Populate the Excel sheet with data from cotizacion
    sheet['D6'] = cotizacion.oficial
    sheet['C10'] = cotizacion.nombreCliente
    sheet['G10'] = cotizacion.cedulaCliente
    # Add more fields as needed
    
    # Save the workbook to a temporary file
    temp_excel_file = os.path.join(settings.BASE_DIR, 'static', 'temp_consultaFideicomiso.xlsx')
    workbook.save(temp_excel_file)
    
    # Convert the Excel file to PDF

    # Load the workbook and select the active sheet
    workbook = load_workbook(temp_excel_file)
    sheet = workbook.active

    # Convert sheet to DataFrame
    data = sheet.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    # Create a PDF document
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Add data to PDF
    for row in dataframe_to_rows(df, index=False, header=True):
        for cell in row:
            pdf.cell(40, 10, str(cell))
        pdf.ln()

    # Save the PDF to a temporary file
    temp_pdf_file = os.path.join(settings.BASE_DIR, 'static', 'temp_consultaFideicomiso.pdf')
    pdf.output(temp_pdf_file)

    return temp_pdf_file


#REPORETE PERSONAL

@login_required
def generate_report_pp(request, numero_cotizacion):
    # Retrieve the cotizacion record based on numero_cotizacion
    print("Generando reporte consulta personalizada")
    try:
        cotizacion = get_object_or_404(Cotizacion, NumeroCotizacion=numero_cotizacion)
        
        # Populate resultado with the values from the cotizacion record
        resultado = {
            'id': 'WEB-' + str(numero_cotizacion),
            'oficial': cotizacion.oficial,
            'sucursal': cotizacion.sucursal,
            'fechaNacimiento': cotizacion.fechaNacimiento,
            'nombreCliente': cotizacion.nombreCliente,
            'cedulaCliente': cotizacion.cedulaCliente,
            'tipoDocumento': cotizacion.tipoDocumento,
            'edad': cotizacion.edad,
            'patrono': cotizacion.patrono,
            'sexo': cotizacion.sexo,
            'apcScore': cotizacion.apcScore,
            'apcPI': cotizacion.apcPI / 100 if cotizacion.apcPI is not None else '',
            'cotPlazoPago': cotizacion.plazoPago,
            'r1': cotizacion.r1,
            'abonoPorcentaje': cotizacion.abonoPorcentaje,
            'abono': cotizacion.abono,
            'cashback': cotizacion.cashback,
            'valorAuto': cotizacion.valorAuto,
            'calcMontoTimbres': cotizacion.calcMontoTimbres,
            'tasaBruta': cotizacion.tasaBruta,
            'aplicaPromocion': cotizacion.aplicaPromocion,
            'cotMontoPrestamo': cotizacion.montoPrestamo,
            'calcMontoNotaria': cotizacion.calcMontoNotaria,
            'promoPublicidad': 50,
            'montoLetraSeguroAdelantado': cotizacion.mesesFinanciaSeguro * cotizacion.montoMensualSeguro,
            'calcComiCierreFinal': cotizacion.calcComiCierreFinal,
            'manejo_5porc': cotizacion.manejo_5porc,
            'auxMonto2': cotizacion.auxMonto2,
            'wrkLetraSinSeguros': cotizacion.wrkLetraSinSeguros,
            'wrkLetraSeguro': cotizacion.wrkLetraSeguro,
            'wrkMontoLetra': cotizacion.wrkMontoLetra,
            'montoMensualSeguro': cotizacion.montoMensualSeguro,
            'wrkLetraConSeguros': cotizacion.wrkMontoLetra + cotizacion.montoMensualSeguro,
            'tablaTotalPagos': cotizacion.tablaTotalPagos,
            'vendedor': cotizacion.vendedor,
            'comisionVendedor': cotizacion.vendedorComision,
            'marcaAuto': cotizacion.marca,
            'lineaAuto': cotizacion.modelo,
            'yearAuto': cotizacion.yearCarro,
            'transmision': cotizacion.transmisionAuto,
            'nuevoAuto': cotizacion.nuevoAuto,
            'kilometrajeAuto': cotizacion.kilometrajeAuto,
            'observaciones': cotizacion.observaciones,
            'salarioBaseMensual': cotizacion.salarioBaseMensual,
            'tiempoServicio': cotizacion.tiempoServicio,
            'ingresos': cotizacion.ingresos,
            'nombreEmpresa': cotizacion.nombreEmpresa,
            'referenciasAPC': cotizacion.referenciasAPC,
            'cartera': cotizacion.cartera,
            'licencia': cotizacion.licencia,
            'posicion': cotizacion.posicion,
            'perfilUniversitario': cotizacion.perfilUniversitario,
            'horasExtrasMonto': cotizacion.horasExtrasMonto,
            'otrosMonto': cotizacion.otrosMonto,
            'montoanualSeguro': cotizacion.montoanualSeguro,
            'otrosDcto': cotizacion.otrosDcto,
            'bonosMonto': cotizacion.bonosMonto,
            'bonosDcto': cotizacion.bonosDcto,
            'siacapMonto': cotizacion.siacapMonto,
            'siacapDcto': cotizacion.siacapDcto,
            'praaMonto': cotizacion.praaMonto,
            'praaDcto': cotizacion.praaDcto,
            'dirOtrosMonto1': cotizacion.dirOtrosMonto1,
            'dirOtros1': cotizacion.dirOtros1,
            'dirOtrosDcto1': cotizacion.dirOtrosDcto1,
            'dirOtrosMonto2': cotizacion.dirOtrosMonto2,
            'dirOtros2': cotizacion.dirOtros2,
            'dirOtrosDcto2': cotizacion.dirOtrosDcto2,
            'dirOtrosMonto3': cotizacion.dirOtrosMonto3,
            'dirOtros3': cotizacion.dirOtros3,
            'dirOtrosDcto3': cotizacion.dirOtrosDcto3,
            'dirOtrosMonto4': cotizacion.dirOtrosMonto4,
            'dirOtros4': cotizacion.dirOtros4,
            'dirOtrosDcto4': cotizacion.dirOtrosDcto4,
            'pagoVoluntario1': cotizacion.pagoVoluntario1,
            'pagoVoluntarioMonto1': cotizacion.pagoVoluntarioMonto1,
            'pagoVoluntarioDcto1': cotizacion.pagoVoluntarioDcto1,
            'pagoVoluntario2': cotizacion.pagoVoluntario2,
            'pagoVoluntarioMonto2': cotizacion.pagoVoluntarioMonto2,
            'pagoVoluntarioDcto2': cotizacion.pagoVoluntarioDcto2,
            'pagoVoluntario3': cotizacion.pagoVoluntario3,
            'pagoVoluntarioMonto3': cotizacion.pagoVoluntarioMonto3,
            'pagoVoluntarioDcto3': cotizacion.pagoVoluntarioDcto3,
            'pagoVoluntario4': cotizacion.pagoVoluntario4,
            'pagoVoluntarioMonto4': cotizacion.pagoVoluntarioMonto4,
            'pagoVoluntarioDcto4': cotizacion.pagoVoluntarioDcto4,
            'pagoVoluntario5': cotizacion.pagoVoluntario5,
            'pagoVoluntarioMonto5': cotizacion.pagoVoluntarioMonto5,
            'pagoVoluntarioDcto5': cotizacion.pagoVoluntarioDcto5,
            'pagoVoluntario6': cotizacion.pagoVoluntario6,
            'pagoVoluntarioMonto6': cotizacion.pagoVoluntarioMonto6,
            'pagoVoluntarioDcto6': cotizacion.pagoVoluntarioDcto6,
            'mes0': cotizacion.mes0,
            'mes1': cotizacion.mes1,
            'mes2': cotizacion.mes2,
            'mes3': cotizacion.mes3,
            'mes4': cotizacion.mes4,
            'mes5': cotizacion.mes5,
            'mes6': cotizacion.mes6,
            'mes7': cotizacion.mes7,
            'mes8': cotizacion.mes8,
            'mes9': cotizacion.mes9,
            'mes10': cotizacion.mes10,
            'mes11': cotizacion.mes11,
            'primerMes': cotizacion.primerMes,
            'tipoProrrateo': cotizacion.tipoProrrateo,
            'tasaInteres': cotizacion.tasaEstimada / 100,
            'tablaTotalSeguro': cotizacion.tablaTotalSeguro,
            'tablaTotalInteres': cotizacion.tablaTotalInteres,
            'tablaTotalFeci': cotizacion.tablaTotalFeci,
            'numeroCotizacion': cotizacion.NumeroCotizacion,
            'montoManejoT': cotizacion.montoManejoT,
            'montoManejoB': cotizacion.monto_manejo_b,
            #codeudor
            'codeudorNombre': cotizacion.codeudorNombre,
            'codeudorCedula': cotizacion.codeudorCedula,
            'codeudorEstabilidad': cotizacion.codeudorEstabilidad,
            'codeudorIngresos': cotizacion.codeudorIngresos,
            'codeudorLicencia': cotizacion.codeudorLicencia,
            'codeudorEmpresa': cotizacion.codeudorEmpresa,
            'codeudorReferenciasAPC': cotizacion.codeudorReferenciasAPC,
            'codeudorNombreEmpres1': cotizacion.codeudorNombreEmpres1,
            'codeudorPeriodo1': cotizacion.codeudorPeriodo1,
            'codeudorSalario1': cotizacion.codeudorSalario1,
            'codeudorNombreEmpres2': cotizacion.codeudorNombreEmpres2,
            'codeudorPeriodo2': cotizacion.codeudorPeriodo2,
            'codeudorSalario2': cotizacion.codeudorSalario2,
            'codeudorNombreEmpres3': cotizacion.codeudorNombreEmpres3,
            'codeudorPeriodo3': cotizacion.codeudorPeriodo3,
            'codeudorSalario3': cotizacion.codeudorSalario3,
            'cobonosMonto': cotizacion.cobonosMonto,
            'cobonosDcto': cotizacion.cobonosDcto,
            'cohorasExtrasMonto': cotizacion.cohorasExtrasMonto,
            'cohorasExtrasDcto': cotizacion.cohorasExtrasDcto,
            'codeudorCartera': cotizacion.codeudorCartera,
            'codeudorPosicion': cotizacion.codeudorPosicion,
            'coprimaMonto': cotizacion.coprimaMonto,
            'coprimaDcto': cotizacion.coprimaDcto,
            'cootrosMonto': cotizacion.cootrosMonto,
            'cootrosDcto': cotizacion.cootrosDcto,
            'cosiacapMonto': cotizacion.cosiacapMonto,
            'cosiacapDcto': cotizacion.cosiacapDcto,
            'copraaMonto': cotizacion.copraaMonto,
            'copraaDcto': cotizacion.copraaDcto,
            'codirOtros1': cotizacion.codirOtros1,
            'codirOtrosMonto1': cotizacion.codirOtrosMonto1,
            'codirOtrosDcto1': cotizacion.codirOtrosDcto1,
            'codirOtros2': cotizacion.codirOtros2,
            'codirOtrosMonto2': cotizacion.codirOtrosMonto2,
            'codirOtrosDcto2': cotizacion.codirOtrosDcto2,
            'codirOtros3': cotizacion.codirOtros3,
            'codirOtrosMonto3': cotizacion.codirOtrosMonto3,
            'codirOtrosDcto3': cotizacion.codirOtrosDcto3,
            'codirOtros4': cotizacion.codirOtros4,
            'codirOtrosMonto4': cotizacion.codirOtrosMonto4,
            'codirOtrosDcto4': cotizacion.codirOtrosDcto4,
            'copagoVoluntario1': cotizacion.copagoVoluntario1,
            'copagoVoluntarioMonto1': cotizacion.copagoVoluntarioMonto1,
            'copagoVoluntarioDcto1': cotizacion.copagoVoluntarioDcto1,
            'copagoVoluntario2': cotizacion.copagoVoluntario2,
            'copagoVoluntarioMonto2': cotizacion.copagoVoluntarioMonto2,
            'copagoVoluntarioDcto2': cotizacion.copagoVoluntarioDcto2,
            'copagoVoluntario3': cotizacion.copagoVoluntario3,
            'copagoVoluntarioMonto3': cotizacion.copagoVoluntarioMonto3,
            'copagoVoluntarioDcto3': cotizacion.copagoVoluntarioDcto3,
            'copagoVoluntario4': cotizacion.copagoVoluntario4,
            'copagoVoluntarioMonto4': cotizacion.copagoVoluntarioMonto4,
            'copagoVoluntarioDcto4': cotizacion.copagoVoluntarioDcto4,
            'copagoVoluntario5': cotizacion.copagoVoluntario5,
            'copagoVoluntarioMonto5': cotizacion.copagoVoluntarioMonto5,
            'copagoVoluntarioDcto5': cotizacion.copagoVoluntarioDcto5,
            'copagoVoluntario6': cotizacion.copagoVoluntario6,
            'copagoVoluntarioMonto6': cotizacion.copagoVoluntarioMonto6,
            'copagoVoluntarioDcto6': cotizacion.copagoVoluntarioDcto6,
            'cancMensualidad1': cotizacion.cancMensualidad1,
            'cancMensualidad2': cotizacion.cancMensualidad2,
            'cancMensualidad3': cotizacion.cancMensualidad3,
            'cancMensualidad4': cotizacion.cancMensualidad4,
            'cancMensualidad5': cotizacion.cancMensualidad5,
            'tapeLetraPrestamo' : cotizacion.tapeLetraPrestamo,




        }
        print('resultado armado')

        
        # Path to the static Excel file
        excel_path = os.path.join(settings.BASE_DIR, 'static/insumos', 'consultaPrestamoPersonal.xlsx')

        if not os.path.exists(excel_path):
            return HttpResponse("File not found.", status=404)
        
        print('cargando excel')
        # Load the workbook and select the active sheet
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        # Select the sheet with name "COTIZADOR PREST. AUTO"
        if "INFORMACIÓN" in workbook.sheetnames:
            sheet = workbook["INFORMACIÓN"]
        else:
            return HttpResponse("Sheet not found.", status=404)
        
        print('llenando excel PP')
        # Example: Write the resultado data to the Excel sheet
        sheet['L7'] = resultado['oficial']
        SUCURSALES_OPCIONES = {
            "2": "COLON",
            "4": "CASINO",
            "7": "DAVID",
            "8": "CHORRERA",
            "11": "SANTIAGO",
            "13": "CALLE 50",
            "16": "CHITRE",
            "17": "PENONOME",
        }
        sheet['D22'] = SUCURSALES_OPCIONES.get(resultado['sucursal'], "DESCONOCIDO")
        sheet['D24'] = resultado['id']
        sheet['D5'] = resultado['nombreCliente']
        sheet['D9'] = resultado['patrono']

        sheet['L5'] = resultado['cartera']
        if resultado['cartera'] == "EMP. PRIVADA":
            sheet['L5'] = "EMP. PRIV."

        sheet['D7'] = resultado['cedulaCliente']
        #sheet['H10'] = resultado['tipoDocumento']
        #sheet['J10'] = resultado['edad']
        #sheet['I10'] = resultado['sexo']
        sheet['D20'] = resultado['fechaNacimiento']
       
        print('Cartera:', resultado['cartera'])
        if resultado['cartera'] in [
            "CONTRALORÍA",
            "AUTÓNOMAS",
            "EMP. PRIVADA",
            "JUBI ACTIVO CONTRALORIA",
            "JUBILADO CONTRALORIA",
            "JUBI ACTIVO AUTÓNOMA",
            "JUBILADO RIESGOS PROF. CSS"
        ]:
            sheet['K60'] = resultado['tapeLetraPrestamo'] * 2
            sheet['d33'] = resultado['tapeLetraPrestamo'] * 2
            sheet['E64'] = resultado['salarioBaseMensual']
            sheet['D29'] = resultado['salarioBaseMensual']
        else:
            sheet['K60'] = resultado['tapeLetraPrestamo'] * 2
            sheet['d33'] = resultado['wrkMontoLetra']
            sheet['E64'] = resultado['salarioBaseMensual'] 
            sheet['D29'] = resultado['salarioBaseMensual'] / 2
            
            
        
        
        #sheet['e42'] = resultado['montoMensualSeguro']
        #sheet['E44'] = resultado['wrkLetraConSeguros']

        #sheet['E46'] = resultado['tablaTotalPagos']
        
        #DATOS DEL VENDEDOR
        #sheet['j18'] = resultado['vendedor']
        #sheet['j20'] = resultado['comisionVendedor']

        #cancelaciones
        sheet['l33'] = resultado['cancMensualidad1']
        sheet['l34'] = resultado['cancMensualidad2']
        sheet['l35'] = resultado['cancMensualidad3']
        sheet['l36'] = resultado['cancMensualidad4']
        sheet['l37'] = resultado['cancMensualidad5']

        #DATOS DEL VEHICULO
       

        #motivo consulta
        
        #sheet['I44'] = resultado['observaciones']

        #promocion
        #if resultado['aplicaPromocion'] == True:
        #     sheet['J42'] = 300
        #else:
        #     sheet['J42'] = ''

        #DATOS DEL DEudor
        
        #sheet['E49']=resultado['tiempoServicio']
        #sheet['J49']=resultado['ingresos']
        
        #sheet['J50']=resultado['referenciasAPC']
        
        #sheet['J51']=resultado['licencia']
        sheet['L9']=resultado['posicion']
        #sheet['E53']=resultado['perfilUniversitario']

        if resultado['horasExtrasMonto'] is None:
            resultado['horasExtrasMonto'] = 0
        if resultado['horasExtrasMonto'] > 0:
            sheet['L65']=resultado['horasExtrasMonto']

        if resultado['otrosMonto'] is None:
            resultado['otrosMonto'] = 0
        if resultado['otrosMonto'] > 0:
            sheet['L68']=resultado['otrosMonto']
            if resultado['otrosDcto'] == True:
                sheet['N68'] = 'SÍ'
            else:
                sheet['N68'] = 'NO'

        if resultado['bonosMonto'] is None:
            resultado['bonosMonto'] = 0
        if resultado['bonosMonto'] > 0:
            sheet['L67']=resultado['bonosMonto']
            if resultado['bonosDcto'] == True:
                sheet['N67'] = 'SÍ'
            else:
                sheet['N67'] = 'NO'
          
        #DESCUENTO DIRECTO
        
        print('siacapMonto', resultado['siacapMonto'])
        if resultado['siacapMonto'] is None:
            resultado['siacapMonto'] = 0
        if resultado['siacapMonto'] > 0:
            sheet['E74'] = resultado['siacapMonto']
            if resultado['siacapDcto'] == True:
                sheet['G74'] = 'SÍ'
            else:
                sheet['G74'] = 'NO'

        if resultado['praaMonto'] is None:
            resultado['praaMonto'] = 0
        if resultado['praaMonto'] > 0:
            sheet['E75'] = resultado['praaMonto']
            if resultado['praaDcto'] == True:
                sheet['G75'] = 'SÍ'
            else:
                sheet['G75'] = 'NO'
        
        if resultado['dirOtrosMonto1'] is None:
            resultado['dirOtrosMonto1'] = 0
        if resultado['dirOtrosMonto1'] > 0:
            sheet['E76'] = resultado['dirOtrosMonto1']
            sheet['C76'] = resultado['dirOtros1']
            if resultado['dirOtrosDcto1'] == True:
                sheet['G76'] = 'SÍ'
            else:
                sheet['G76'] = 'NO'

        if resultado['dirOtrosMonto2'] is None:
            resultado['dirOtrosMonto2'] = 0

        if resultado['dirOtrosMonto2'] > 0:
            sheet['E77'] = resultado['dirOtrosMonto2']
            sheet['C77'] = resultado['dirOtros2']
            if resultado['dirOtrosDcto2'] == True:
                sheet['G77'] = 'SÍ'
            else:
                sheet['G77'] = 'NO'

        if resultado['dirOtrosMonto3'] is None:
            resultado['dirOtrosMonto3'] = 0

        if resultado['dirOtrosMonto3'] > 0:
            sheet['E78'] = resultado['dirOtrosMonto3']
            sheet['c78'] = resultado['dirOtros3']
            if resultado['dirOtrosDcto3'] == True:
                sheet['G78'] = 'SÍ'
            else:
                sheet['G78'] = 'NO'

        if resultado['dirOtrosMonto4'] is None:
            resultado['dirOtrosMonto4'] = 0


        if resultado['dirOtrosMonto4'] > 0:
            sheet['e79'] = resultado['dirOtrosMonto4']
            sheet['C79'] = resultado['dirOtros4']
            if resultado['dirOtrosDcto4'] == True:
                sheet['g79'] = 'SÍ'
            else:
                sheet['g79'] = 'NO'

        #pagos voluntarios
        sheet['I74'] = resultado['pagoVoluntario1']
        sheet['I75'] = resultado['pagoVoluntario2']
        sheet['I76'] = resultado['pagoVoluntario3']
        
        sheet['I77'] = resultado['pagoVoluntario4']
        
        sheet['I78'] = resultado['pagoVoluntario5']
        
        sheet['I79'] = resultado['pagoVoluntario6']
        
        

        #PARSE TRUE TO SI AND FALSE TO NO
        if resultado ['pagoVoluntarioMonto1'] is None:
            resultado['pagoVoluntarioMonto1'] = 0

        if resultado['pagoVoluntarioMonto1'] > 0:
            sheet['M74'] = resultado['pagoVoluntarioMonto1']
            if resultado['pagoVoluntarioDcto1'] == True:
                sheet['O74'] = 'SÍ'
            else:
                sheet['O74'] = 'NO'

        if resultado['pagoVoluntarioMonto2'] is None:
            resultado['pagoVoluntarioMonto2'] = 0
        if resultado['pagoVoluntarioMonto2'] > 0:
            sheet['M75'] = resultado['pagoVoluntarioMonto2']
            if resultado['pagoVoluntarioDcto2'] == True:
                sheet['O75'] = 'SÍ'
            else:
                sheet['O75'] = 'NO'

        if resultado['pagoVoluntarioMonto3'] is None:
            resultado['pagoVoluntarioMonto3'] = 0
        if resultado['pagoVoluntarioMonto3'] > 0:
            sheet['M76'] = resultado['pagoVoluntarioMonto3']
            if resultado['pagoVoluntarioDcto3'] == True:
                sheet['O76'] = 'SÍ'
            else:
                sheet['O76'] = 'NO'
        if resultado['pagoVoluntarioMonto4'] is None:
            resultado['pagoVoluntarioMonto4'] = 0

        if resultado['pagoVoluntarioMonto4'] > 0:
            sheet['M77'] = resultado['pagoVoluntarioMonto4']
            if resultado['pagoVoluntarioDcto4'] == True:
                sheet['O77'] = 'SÍ'
            else:
                sheet['O77'] = 'NO'

        if resultado['pagoVoluntarioMonto5'] is None:
            resultado['pagoVoluntarioMonto5'] = 0

        if resultado['pagoVoluntarioMonto5'] > 0:
            sheet['M78'] = resultado['pagoVoluntarioMonto5']
            if resultado['pagoVoluntarioDcto5'] == True:
                sheet['O78'] = 'SÍ'
            else:
                sheet['O78'] = 'NO'
        
        if resultado['pagoVoluntarioMonto6'] is None:
            resultado['pagoVoluntarioMonto6'] = 0

        if resultado['pagoVoluntarioMonto6'] > 0:
            sheet['M79'] = resultado['pagoVoluntarioMonto6']
            if resultado['pagoVoluntarioDcto6'] == True:
                sheet['O79'] = 'SÍ'
            else:
                sheet['O79'] = 'NO'

        #Codeudor
        sheet['D89'] = resultado['codeudorNombre']
        sheet['D91'] = resultado['codeudorCedula']
        sheet['L89'] = resultado['codeudorCartera']
        sheet['L93'] = resultado['codeudorPosicion']

        #sheet['E56'] = resultado['codeudorEstabilidad']
        sheet['E99'] = resultado['codeudorIngresos']
        #sheet['J57'] = resultado['codeudorLicencia']
        sheet['L91'] = resultado['codeudorEmpresa']
       

        #CODEUDOR - INGRESOS ADICIONALES
        if resultado['cohorasExtrasMonto'] is None:
            resultado['cohorasExtrasMonto'] = 0
        if resultado['cohorasExtrasMonto'] > 0:
            sheet['L100']=resultado['cohorasExtrasMonto']
            if resultado['cohorasExtrasDcto'] == True:
                sheet['N100'] = 'SÍ'
            else:
                sheet['N100'] = 'NO'

        if resultado['coprimaMonto'] is None:
            resultado['coprimaMonto'] = 0
        if resultado['coprimaMonto'] > 0:
            sheet['L101']=resultado['coprimaMonto']
            if resultado['coprimaDcto'] == True:
                sheet['N101'] = 'SÍ'
            else:
                sheet['N101'] = 'NO'

        if resultado['cobonosMonto'] is None:
            resultado['cobonosMonto'] = 0
        if resultado['cobonosMonto'] > 0:
            sheet['L102']=resultado['cobonosMonto']
            if resultado['cobonosDcto'] == True:
                sheet['N102'] = 'SÍ'
            else:
                sheet['N102'] = 'NO'

        if resultado['cootrosMonto'] is None:
            resultado['cootrosMonto'] = 0
        if resultado['cootrosMonto'] > 0:
            sheet['L103']=resultado['cootrosMonto']
            if resultado['cootrosDcto'] == True:
                sheet['N103'] = 'SÍ'
            else:
                sheet['N103'] = 'NO'
    #codeudor - DEscuentos directos
        if resultado['cosiacapMonto'] is None:
                resultado['cosiacapMonto'] = 0
        if resultado['cosiacapMonto'] > 0:
                sheet['E109']=resultado['cosiacapMonto']
                if resultado['cosiacapDcto'] == True:
                    sheet['G109'] = 'SÍ'
                else:
                    sheet['G109'] = 'NO'

        if resultado['copraaMonto'] is None:
                resultado['copraaMonto'] = 0
        if resultado['copraaMonto'] > 0:
                sheet['E110']=resultado['copraaMonto']
                if resultado['copraaDcto'] == True:
                    sheet['G110'] = 'SÍ'
                else:
                    sheet['G110'] = 'NO'

        if resultado['codirOtrosMonto1'] is None:
                resultado['codirOtrosMonto1'] = 0
        if resultado['codirOtrosMonto1'] > 0:
                sheet['E111']=resultado['codirOtrosMonto1']
                sheet['C111']=resultado['codirOtros1']
                if resultado['codirOtrosDcto1'] == True:
                    sheet['G111'] = 'SÍ'
                else:
                    sheet['G111'] = 'NO'
        
        if resultado['codirOtrosMonto2'] is None:
                resultado['codirOtrosMonto2'] = 0
        if resultado['codirOtrosMonto2'] > 0:
                sheet['E112']=resultado['codirOtrosMonto2']
                sheet['C112']=resultado['codirOtros2']
                if resultado['codirOtrosDcto2'] == True:
                    sheet['G112'] = 'SÍ'
                else:
                    sheet['G112'] = 'NO'

        if resultado['codirOtrosMonto3'] is None:
                resultado['codirOtrosMonto3'] = 0
        if resultado['codirOtrosMonto3'] > 0:
                sheet['E113']=resultado['codirOtrosMonto3']
                sheet['C113']=resultado['codirOtros3']
                if resultado['codirOtrosDcto3'] == True:
                    sheet['G113'] = 'SÍ'
                else:
                    sheet['G113'] = 'NO'

        if resultado['codirOtrosMonto4'] is None:
                resultado['codirOtrosMonto4'] = 0
        if resultado['codirOtrosMonto4'] > 0:
                sheet['E114']=resultado['codirOtrosMonto4']
                sheet['C114']=resultado['codirOtros4']
                if resultado['codirOtrosDcto4'] == True:
                    sheet['G114'] = 'SÍ'
                else:
                    sheet['G114'] = 'NO'
    #CODEUDOR - PAGOS VOLUNTARIOS
        if resultado['copagoVoluntarioMonto1'] is None:
                resultado['copagoVoluntarioMonto1'] = 0
        if resultado['copagoVoluntarioMonto1'] > 0:
                sheet['M109']=resultado['copagoVoluntarioMonto1']
                sheet['I109']=resultado['copagoVoluntario1']
                if resultado['copagoVoluntarioDcto1'] == True:
                    sheet['O109'] = 'SÍ'
                else:
                    sheet['O109'] = 'NO'
        
        if resultado['copagoVoluntarioMonto2'] is None:
                resultado['copagoVoluntarioMonto2'] = 0
        if resultado['copagoVoluntarioMonto2'] > 0:
                sheet['M110']=resultado['copagoVoluntarioMonto2']
                sheet['I110']=resultado['copagoVoluntario2']
                if resultado['copagoVoluntarioDcto2'] == True:
                    sheet['O110'] = 'SÍ'
                else:
                    sheet['O110'] = 'NO'
        
        if resultado['copagoVoluntarioMonto3'] is None:
                resultado['copagoVoluntarioMonto3'] = 0
        if resultado['copagoVoluntarioMonto3'] > 0:
                sheet['M111']=resultado['copagoVoluntarioMonto3']
                sheet['I111']=resultado['copagoVoluntario3']
                if resultado['copagoVoluntarioDcto3'] == True:
                    sheet['O111'] = 'SÍ'
                else:
                    sheet['O111'] = 'NO'

        if resultado['copagoVoluntarioMonto4'] is None:
                resultado['copagoVoluntarioMonto4'] = 0
        if resultado['copagoVoluntarioMonto4'] > 0:
                sheet['M112']=resultado['copagoVoluntarioMonto4']
                sheet['I112']=resultado['copagoVoluntario4']
                if resultado['copagoVoluntarioDcto4'] == True:
                    sheet['O112'] = 'SÍ'
                else:
                    sheet['O112'] = 'NO'
        
        if resultado['copagoVoluntarioMonto5'] is None:
                resultado['copagoVoluntarioMonto5'] = 0
        if resultado['copagoVoluntarioMonto5'] > 0: 
                sheet['M113']=resultado['copagoVoluntarioMonto5']
                sheet['I113']=resultado['copagoVoluntario5']
                if resultado['copagoVoluntarioDcto5'] == True:
                    sheet['O113'] = 'SÍ'
                else:
                    sheet['O113'] = 'NO'

        if resultado['copagoVoluntarioMonto6'] is None: 
                resultado['copagoVoluntarioMonto6'] = 0
        if resultado['copagoVoluntarioMonto6'] > 0:
                sheet['M114']=resultado['copagoVoluntarioMonto6']
                sheet['I114']=resultado['copagoVoluntario6']
                if resultado['copagoVoluntarioDcto6'] == True:
                    sheet['O114'] = 'SÍ'
                else:
                    sheet['O114'] = 'NO'

        print('llenando excel')
     # Select the sheet with name "PRORRATEO"
        if "PRORRATEO" in workbook.sheetnames:
            prorrateo = workbook["PRORRATEO"]
        else:
            return HttpResponse("Sheet not found.", status=404)
        
        prorrateo['D6'] = resultado['mes0']
        prorrateo['d7'] = resultado['mes1']
        prorrateo['d8'] = resultado['mes2']
        prorrateo['d9'] = resultado['mes3']
        prorrateo['d10'] = resultado['mes4']
        prorrateo['d11'] = resultado['mes5']
        prorrateo['d12'] = resultado['mes6']
        prorrateo['d13'] = resultado['mes7']
        prorrateo['d14'] = resultado['mes8']
        prorrateo['d15'] = resultado['mes9']
        prorrateo['d16'] = resultado['mes10']
        prorrateo['d17'] = resultado['mes11']
        prorrateo['C6'] = resultado['primerMes']
        print('prmer mes', resultado['primerMes'])

         # Select the sheet with name "DESGLOSE"
        

        # Select the sheet with name "mov bancarios"
        if "MOV. BANCARIOS" in workbook.sheetnames:
            movimientos = workbook["MOV. BANCARIOS"]
        else:
            return HttpResponse("Sheet not found. Mov", status=404)
        
        movimientos['C6'] = cotizacion.movPrimerMes if cotizacion.movPrimerMes is not None else "ENERO"
        movimientos['C8'] = cotizacion.ingresosMes1 if cotizacion.ingresosMes1 is not None else 0
        movimientos['D8'] = cotizacion.egresosMes1 if cotizacion.egresosMes1 is not None else 0
        movimientos['f8'] = cotizacion.ingresosMes2 if cotizacion.ingresosMes2 is not None else 0
        movimientos['g8'] = cotizacion.egresosMes2 if cotizacion.egresosMes2 is not None else 0
        movimientos['i8'] = cotizacion.ingresosMes3 if cotizacion.ingresosMes3 is not None else 0
        movimientos['j8'] = cotizacion.egresosMes3 if cotizacion.egresosMes3 is not None else 0
        movimientos['l8'] = cotizacion.ingresosMes4 if cotizacion.ingresosMes4 is not None else 0
        movimientos['m8'] = cotizacion.egresosMes4 if cotizacion.egresosMes4 is not None else 0
        movimientos['o8'] = cotizacion.ingresosMes5 if cotizacion.ingresosMes5 is not None else 0
        movimientos['p8'] = cotizacion.egresosMes5 if cotizacion.egresosMes5 is not None else 0
        movimientos['r8'] = cotizacion.ingresosMes6 if cotizacion.ingresosMes6 is not None else 0
        movimientos['s8'] = cotizacion.egresosMes6 if cotizacion.egresosMes6 is not None else 0
        try:
            if cotizacion.movOpcion == "tomar_valor":
                movimientos['U13'] = "TOMAR VALOR"
            else:
                movimientos['U13'] = cotizacion.movOpcion if cotizacion.movOpcion is not None else 'COLOCAR MANUAL'
        except:
                movimientos['U13'] = "COLOCAR MANUAL"

        
        #LLENADO RESUMEN
        sheet = workbook['RESUMEN']
        sheet['I5'] = resultado['id']
        sheet['K5'] = resultado['oficial']
        sheet['K5'] = resultado['oficial']

        sheet['E11'] = cotizacion.fechaInicioPago
        sheet['E12'] = cotizacion.fechaVencimiento
        if cotizacion.formaPago == 1:
            sheet['E9'] = "PAGO VOLUNTARIO"
        elif cotizacion.formaPago == 2:
            sheet['E9'] = "DESCUENTO DIRECTO"

        sheet['c7'] = cotizacion.r1
        sheet['c9'] = cotizacion.tasaEstimada
        if cotizacion.periodoPago == 1:
            sheet['C10'] = "MENSUAL"
        elif cotizacion.periodoPago == 2:
            sheet['C10'] = "QUINCENAL" 
            
        sheet['c11'] = "SOBRESALDO"
        sheet['C12'] = cotizacion.tasaBruta

        sheet['c14'] = cotizacion.auxMonto2

        sheet['c16'] = cotizacion.calcMontoTimbres
        sheet['c17'] = cotizacion.monto_manejo_b
        sheet['c18'] = cotizacion.montoServDesc
        sheet['c19'] = cotizacion.manejo_5porc
        sheet['c20'] = cotizacion.montoManejoT
        sheet['H10'] = str(cotizacion.plazoPago) + " / " + str(cotizacion.plazoInteres)
        sheet['H11'] = cotizacion.pagaDiciembre
        sheet['h15'] = cotizacion.montoPrestamo 
        sheet['H16'] = cotizacion.tablaTotalPagos  
        sheet['H14'] = cotizacion.calcMontoNotaria

        sheet['H23'] = cotizacion.vendedor
        sheet['H24'] = cotizacion.vendedorComisionPorcentaje

        if cotizacion.periodoPago == 1:
             sheet['H19'] = cotizacion.plazoPago
             sheet['H20'] = cotizacion.wrkMontoLetra
        elif cotizacion.periodoPago == 2:
            sheet['H19'] = cotizacion.plazoPago * 2
            sheet['H20'] = cotizacion.wrkMontoLetra * 2

        print('llenando excel')
        # Save the workbook to a temporary file
        temp_file = os.path.join(settings.BASE_DIR, 'static', 'temp_consultaFideicomiso.xlsx')
        workbook.save(temp_file)
        
        # Serve the file as a response
        nombre_cliente = resultado['nombreCliente']
        filename = f"Consulta - {numero_cotizacion} -{nombre_cliente}.xlsx"
        with open(temp_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response
    
    except Exception as e:
        error_message = str(e)
        # Log the error
        print(f"Error: {error_message}, User: {request.user.username}")
        return JsonResponse({'status': 'error', 'message': str(e), 'resultado': resultado}, status=500)
  
